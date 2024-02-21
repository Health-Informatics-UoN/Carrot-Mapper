import asyncio
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Literal, Optional

import azure.functions as func
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from shared_code.api import (
    ScanReportStatus,
    get_concept_vocabs,
    get_scan_report_active_concepts,
    get_scan_report_fields_by_ids,
    get_scan_report_values,
    get_scan_report_values_filter_scan_report_table,
    post_chunks,
    post_scan_report_concepts,
    update_scan_report_status,
)
from shared_code.logger import logger

from shared_code import blob_parser, helpers, omop_helpers

# import memory_profiler
# root_logger = logging.getLogger()
# root_logger.handlers[0].setFormatter(logging.Formatter("%(name)s: %(message)s"))
# profiler_logstream = memory_profiler.LogFile('memory_profiler_logs', True)


def select_concepts_to_post(
    new_content_details: List[Dict[Optional[str], Optional[str]]],
    details_to_id_and_concept_id_map: List[Dict[str, str]],
    content_type: Literal[15, 17],
) -> List[Dict[str, Any]]:
    """
    Depending on the content_type, generate a list of `ScanReportConcepts` to be created.
    Content_type controls whether this is handling fields or values.
    Fields have a key defined only by name, while values have a key defined by:
      name, description, and field name.

    Args:
        new_content_details (List[Dict[str, str, str (optional), str (optional)]]):
          Each item in the list a dict containing either "id" and "name" keys (for fields)
          or "id", "name", "description", and "field_name" keys (for values).

        details_to_id_and_concept_id_map (List[Dict[str, str, str]]): keys "name" (for fields) or ("name",
          "description", "field_name") keys (for values), with entries (field_id, concept_id)
          or (value_id, concept_id) respectively.

        content_type (Literal[15, 17]): Controls whether to handle fields (15), or values (17).

    Returns:
        A list of `Concepts` that look like `ScanReportConcepts` to create.

    Raises:
        Exception: RunTimeError: A content_type other than 15 or 17 was provided.
    """
    concepts_to_post = []
    for new_content_detail in new_content_details:
        try:
            # fields
            if content_type == 15:
                (
                    existing_content_id,
                    concept_id,
                ) = details_to_id_and_concept_id_map[str(new_content_detail["name"])]
            # values
            elif content_type == 17:
                existing_content_id, concept_id = details_to_id_and_concept_id_map[
                    (
                        str(new_content_detail["name"]),
                        str(new_content_detail["description"]),
                        str(new_content_detail["field_name"]),
                    )
                ]
            else:
                raise RuntimeError(
                    f"content_type must be 15 or 17: you provided {content_type}"
                )

            new_content_id = str(new_content_detail["id"])
            if content_type == 15:
                logger.info(
                    f"Found existing field with id: {existing_content_id} with existing "
                    f"concept mapping: {concept_id} which matches new field id: "
                    f"{new_content_id}"
                )
            elif content_type == 17:
                logger.info(
                    f"Found existing value with id: {existing_content_id} with existing "
                    f"concept mapping: {concept_id} which matches new value id: "
                    f"{new_content_id}"
                )
            # Create ScanReportConcept entry for copying over the concept
            concept_entry = {
                "nlp_entity": None,
                "nlp_entity_type": None,
                "nlp_confidence": None,
                "nlp_vocabulary": None,
                "nlp_processed_string": None,
                "concept": concept_id,
                "object_id": new_content_id,
                "content_type": content_type,
                "creation_type": "R",
            }
            concepts_to_post.append(concept_entry)
        except KeyError:
            continue

    return concepts_to_post


def reuse_existing_field_concepts(
    new_fields_map: Dict[str, str], content_type: Literal[15]
) -> None:
    """
    It creates new concepts associated to any field that matches the name of an existing
    field with an associated concept.

    This expects a dict of field names to ids which have been generated in a newly uploaded
    scanreport, and content_type 15.

    Args:
        new_fields_map (Dict[str, str]): A map of field names to Ids.
        content_type (Literal[15]): The content type, represents `ScanReportField` (15)
    """
    logger.info("reuse_existing_field_concepts")
    # Gets all scan report concepts that are for the type field
    # (or content type which should be field) and in "active" SRs
    existing_field_concepts = get_scan_report_active_concepts(content_type)

    # create dictionary that maps existing field ids to scan report concepts
    # from the list of existing scan report concepts from active SRs
    existing_field_id_to_concept_map = {
        str(element.get("object_id", None)): str(element.get("concept", None))
        for element in existing_field_concepts
    }
    logger.debug(
        f"field_id:concept_id for all existing fields in active SRs with concepts: "
        f"{existing_field_id_to_concept_map}"
    )

    # get details of existing selected fields, for the purpose of matching against
    # new fields
    existing_field_ids = {item["object_id"] for item in existing_field_concepts}
    existing_fields = get_scan_report_fields_by_ids(existing_field_ids)
    logger.debug(
        f"ids and names of existing fields with concepts in active SRs:"
        f" {existing_fields}"
    )

    # Combine everything of the existing fields to get this list, one for each
    # existing SRField with a SRConcept in an active SR.
    existing_mappings_to_consider = [
        {
            "name": field["name"],
            "concept": existing_field_id_to_concept_map[str(field["id"])],
            "id": field["id"],
        }
        for field in existing_fields
    ]
    logger.debug(f"{existing_mappings_to_consider=}")

    # Handle the newly-added fields
    new_fields_full_details = [
        {"name": name, "id": new_field_id}
        for name, new_field_id in new_fields_map.items()
    ]
    # Now we have existing_mappings_to_consider of the form:
    #
    # [{"id":, "name":, "concept":}]
    #
    # and new_fields_full_details of the form:
    #
    # [{"id":, "name":}]

    # Now we simply look for unique matches on "name" across
    # the two.

    # existing_field_name_to_field_and_concept_id_map will contain
    # (field_name) -> (field_id, concept_id)
    # for each field in new_fields_full_details
    # if that field has only one match in existing_mappings_to_consider
    existing_field_name_to_field_and_concept_id_map = {}
    for item in new_fields_full_details:
        name = item["name"]
        mappings_matching_field_name = list(
            filter(
                lambda mapping: mapping["name"] == name, existing_mappings_to_consider
            )
        )

        target_concept_ids = {
            mapping["concept"] for mapping in mappings_matching_field_name
        }

        if len(target_concept_ids) == 1:
            target_field_id = (
                mapping["id"] for mapping in mappings_matching_field_name
            )
            existing_field_name_to_field_and_concept_id_map[str(name)] = (
                str(next(target_field_id)),
                str(target_concept_ids.pop()),
            )

    # Use the new_fields_full_details as keys into
    # existing_field_name_to_field_and_concept_id_map to extract concept IDs and details
    # for new ScanReportConcept entries to post.
    if concepts_to_post := select_concepts_to_post(
        new_fields_full_details,
        existing_field_name_to_field_and_concept_id_map,
        content_type,
    ):
        post_scan_report_concepts(concepts_to_post)
        logger.info("POST concepts all finished in reuse_existing_field_concepts")


def reuse_existing_value_concepts(new_values_map, content_type: Literal[17]) -> None:
    """
    This expects a dict of value names to ids which have been generated in a newly
    uploaded scanreport and creates new concepts if any matching names are found
    with existing fields

    TODO: Why the f is this is parameter if it only accepts one specific number?
    Args:
        new_fields_map (Dict[str, str]): A map of field names to Ids.
        content_type (Literal[17]): The content type, represents `ScanReportValue` (17)
    """
    logger.info("reuse_existing_value_concepts")
    # Gets all scan report concepts that are for the type value
    # (or content type which should be value) and in "active" SRs

    existing_value_concepts = get_scan_report_active_concepts(content_type)
    # create dictionary that maps existing value ids to scan report concepts
    # from the list of existing scan report concepts
    existing_value_id_to_concept_map = {
        str(element.get("object_id", None)): str(element.get("concept", None))
        for element in existing_value_concepts
    }
    logger.debug(
        f"value_id:concept_id for all existing values in active SRs with concepts: "
        f"{existing_value_id_to_concept_map}"
    )

    # get details of existing selected values, for the purpose of matching against
    # new values
    existing_paginated_value_ids = helpers.paginate(
        [value["object_id"] for value in existing_value_concepts],
        omop_helpers.max_chars_for_get,
    )
    logger.debug(f"{existing_paginated_value_ids=}")

    # for each list in paginated ids, get scanreport values that match any of the given
    # ids (those with an associated concept)
    existing_values_filtered_by_id = []
    for ids in existing_paginated_value_ids:
        ids_to_get = ",".join(map(str, ids))
        values = get_scan_report_values(ids_to_get)
        existing_values_filtered_by_id.append(values)
    existing_values_filtered_by_id = helpers.flatten(existing_values_filtered_by_id)
    logger.debug("existing_values_filtered_by_id")

    # existing_values_filtered_by_id now contains the id,value,value_dec,
    # scan_report_field of each value got from the active concepts filter.
    logger.debug(
        f"ids, values and value descriptions of existing values in active SRs with "
        f"concepts: {existing_values_filtered_by_id}"
    )

    # get field ids from values and use to get scan report fields' details
    existing_field_ids = {
        item["scan_report_field"] for item in existing_values_filtered_by_id
    }
    existing_fields = get_scan_report_fields_by_ids(existing_field_ids)
    logger.debug(
        f"ids and names of existing fields associated to values with concepts in "
        f"active SRs: {existing_fields}"
    )

    existing_field_id_to_name_map = {
        str(field["id"]): field["name"] for field in existing_fields
    }
    logger.debug(f"{existing_field_id_to_name_map=}")

    # Combine everything of the existing values to get this list, one for each
    # existing SRValue with a SRConcept in an active SR.
    existing_mappings_to_consider = [
        {
            "name": value["value"],
            "concept": existing_value_id_to_concept_map[str(value["id"])],
            "id": value["id"],
            "description": value["value_description"],
            "field_name": existing_field_id_to_name_map[
                str(value["scan_report_field"])
            ],
        }
        for value in existing_values_filtered_by_id
    ]

    # Now handle the newly-added values in a similar manner
    logger.debug("new_paginated_field_ids")
    new_field_ids = [value["scan_report_field"] for value in new_values_map]
    new_fields = get_scan_report_fields_by_ids(new_field_ids)
    logger.debug(f"fields of newly generated values: {new_fields}")

    new_fields_to_name_map = {str(field["id"]): field["name"] for field in new_fields}

    new_values_full_details = [
        {
            "name": value["value"],
            "description": value["value_description"],
            "field_name": new_fields_to_name_map[str(value["scan_report_field"])],
            "id": value["id"],
        }
        for value in new_values_map
    ]

    # Now we have existing_mappings_to_consider of the form:
    #
    # [{"id":, "name":, "concept":, "description":, "field_name":}]
    #
    # and new_values_full_details of the form:
    #
    # [{"id":, "name":, "description":, "field_name":}]

    # Now we simply look for unique matches on (name, description, field_name) across
    # the two.

    # value_details_to_value_and_concept_id_map will contain
    # (name, description, field_name) -> (value_id, concept_id)
    # for each value/description/field tuple in new_values_full_details
    # if that tuple has only one match in existing_mappings_to_consider
    value_details_to_value_and_concept_id_map = {}
    for item in new_values_full_details:
        name = item["name"]
        description = item["description"]
        field_name = item["field_name"]
        mappings_matching_value_name = list(
            filter(
                lambda mapping: mapping["name"] == name
                and mapping["description"] == description
                and mapping["field_name"] == field_name,
                existing_mappings_to_consider,
            )
        )

        target_concept_ids = {
            mapping["concept"] for mapping in mappings_matching_value_name
        }

        if len(target_concept_ids) == 1:
            target_value_id = (
                mapping["id"] for mapping in mappings_matching_value_name
            )
            value_details_to_value_and_concept_id_map[
                (str(name), str(description), str(field_name))
            ] = (str(next(target_value_id)), str(target_concept_ids.pop()))

    # Use the new_values_full_details as keys into
    # value_details_to_value_and_concept_id_map to extract concept IDs and details
    # for new ScanReportConcept entries to post.
    if concepts_to_post := select_concepts_to_post(
        new_values_full_details,
        value_details_to_value_and_concept_id_map,
        content_type,
    ):
        post_scan_report_concepts(concepts_to_post)
        logger.info("POST concepts all finished in reuse_existing_value_concepts")


# @memory_profiler.profile(stream=profiler_logstream)
async def process_values_for_table(
    vocab_dictionary: Dict[str, Any],
    table_name: str,
    table_id: str,
    fieldnames_to_ids_dict: Dict[str, str],
    fieldids_to_names_dict: Dict[str, str],
    scan_report_id: str,
) -> None:
    """
    This function handles much of the complexity surrounding values for a given table.
    They can have a value description from the data dictionary, and they can have a
    vocab mapping from vocab dictionary.

    In summary, we:
    - get the details of the existing SRValues
    - apply vocab mapping to each SRValue as appropriate. Much of the complexity and
    audit-keeping is because of the requirement to do this with batch calls - single
    calls are simply too slow.

    Args:
        vocab_dictionary (Dict[str, Any]):
        table_name: (str): Name of the table
        table_id (str): Id of the table
        fieldnames_to_ids_dict (Dict[str, str]): ???
        fieldids_to_names_dict (Dict[str, str]): ???
        scan_report_id: Id of the Scan Report.
    """
    # TODO: This is where we split.
    # We want to remove as many parameters as possible to this function.

    # --------------------------------------------------------------------------------
    # Get the details of all the SRValues posted in this table. Then we will be able
    # to run them through the vocabulary mapper and apply any automatic vocab mappings.

    # GET values where the scan_report_table is the current table.
    logger.debug("GET existing values")
    details_of_posted_values = get_scan_report_values_filter_scan_report_table(table_id)
    logger.debug("GET existing values finished")

    # ---------------------------------------------------------------------------------
    # Process the SRValues, comparing their SRFields to the vocabs, and then create a
    # SRConcept entry if a valid translation is found.

    # ------------------------------------------------------------------------------
    # Add vocabulary_id to each entry from the vocab dictionary
    helpers.add_vocabulary_id_to_entries(
        details_of_posted_values, vocab_dictionary, fieldids_to_names_dict, table_name
    )
    logger.debug("split by vocab")

    # ---------------------------------------------------------------------
    # Split up by each vocabulary_id, so that we can send the contents of each to the
    # endpoint in a separate call, with one vocabulary_id per call. Note that the
    # entries in entries_split_by_vocab are views of those in
    # details_of_posted_values, which is the variable we will eventually use to
    # populate our SRConcept entries. So which we mostly operate on
    # entries_split_by_vocab[vocab[ below, these changes are accessible from
    # details_of_posted_values
    entries_split_by_vocab = defaultdict(list)
    for entry in details_of_posted_values:
        entries_split_by_vocab[entry["vocabulary_id"]].append(entry)

    # ----------------------------------------------
    # For each vocab, set "concept_id" and "standard_concept" in each entry in the
    # vocab.
    #
    # For the case when vocab is None, set it to defaults.
    #
    # For other cases, get the concepts from the vocab via /omop/conceptsfilter under
    # pagination.
    # Then match these back to the originating values, setting "concept_id" and
    # "standard_concept" in each case.
    # Finally, we need to fix all entries where "standard_concept" != "S" using
    # `find_standard_concept_batch()`. This may result in more than one standard
    # concept for a single nonstandard concept, and so "concept_id" may be either an
    # int or str, or a list of such.

    for vocab, value in entries_split_by_vocab.items():
        if vocab is None:
            # set to defaults, and skip all the remaining processing that a vocab
            # would require
            for entry in value:
                entry["concept_id"] = -1
                entry["standard_concept"] = None
            continue

        assert vocab is not None
        logger.info(f"begin {vocab}")

        paginated_values_in_this_vocab = helpers.paginate(
            (str(entry["value"]) for entry in entries_split_by_vocab[vocab]),
            max_chars=omop_helpers.max_chars_for_get,
        )

        concept_vocab_response = []

        for page_of_values in paginated_values_in_this_vocab:
            page_of_values_to_get = ",".join(map(str, page_of_values))

            concept_vocab = get_concept_vocabs(vocab, page_of_values_to_get)
            concept_vocab_response.append(concept_vocab)

        concept_vocab_content = helpers.flatten(concept_vocab_response)

        # Loop over all returned concepts, and match their concept_code and vocabulary_id with
        # the full_value in the entries_split_by_vocab, and set the latter's
        # concept_id and standard_concept with those values
        logger.debug(
            f"Attempting to match {len(concept_vocab_content)} concepts to "
            f"{len(entries_split_by_vocab[vocab])} SRValues"
        )
        for entry in entries_split_by_vocab[vocab]:
            entry["concept_id"] = -1
            entry["standard_concept"] = None

        for entry in entries_split_by_vocab[vocab]:
            for returned_concept in concept_vocab_content:
                if str(entry["value"]) == str(returned_concept["concept_code"]):
                    entry["concept_id"] = str(returned_concept["concept_id"])
                    entry["standard_concept"] = str(
                        returned_concept["standard_concept"]
                    )
                    # exit inner loop early if we find a concept for this entry
                    break

        logger.debug("finished matching")

        # ------------------------------------------------
        # Identify which concepts are non-standard, and get their standard counterparts
        # in a batch call
        entries_to_find_standard_concept = list(
            filter(
                lambda x: x["concept_id"] != -1 and x["standard_concept"] != "S",
                entries_split_by_vocab[vocab],
            )
        )
        logger.debug(
            f"finished selecting nonstandard concepts - selected "
            f"{len(entries_to_find_standard_concept)}"
        )

        batched_standard_concepts_map = omop_helpers.find_standard_concept_batch(
            entries_to_find_standard_concept
        )

        # batched_standard_concepts_map maps from an original concept id to
        # a list of associated standard concepts. Use each item to update the
        # relevant entry from entries_split_by_vocab[vocab].
        for nonstandard_concept in batched_standard_concepts_map:
            relevant_entry = helpers.get_by_concept_id(
                entries_split_by_vocab[vocab], nonstandard_concept
            )

            if isinstance(relevant_entry["concept_id"], (int, str)):
                relevant_entry["concept_id"] = batched_standard_concepts_map[
                    nonstandard_concept
                ]
            elif relevant_entry["concept_id"] is None:
                # This is the case where pairs_for_use contains an entry that
                # doesn't have a counterpart in entries_split_by_vocab, so this
                # should error or warn
                raise RuntimeWarning

        logger.debug("finished standard concepts lookup")

    # ------------------------------------
    # All Concepts are now ready. Generate their entries ready for POSTing from
    # details_of_posted_values. Remember that entries_split_by_vocab is just a view
    # into this list, so changes to entries_split_by_vocab above are reflected when
    # we access details_of_posted_values below.

    concept_id_data = []
    for concept in details_of_posted_values:
        if concept["concept_id"] != -1:
            if isinstance(concept["concept_id"], list):
                concept_id_data.extend(
                    {
                        "concept": concept_id,
                        "object_id": concept["id"],
                        # TODO: we should query this value from the API
                        # - via ORM it would be ContentType.objects.get(model='scanreportvalue').id,
                        # but that's not available from an Azure Function.
                        "content_type": 17,
                        "creation_type": "V",
                    }
                    for concept_id in concept["concept_id"]
                )
            else:
                concept_id_data.append(
                    {
                        "concept": concept["concept_id"],
                        "object_id": concept["id"],
                        # TODO: we should query this value from the API
                        # - via ORM it would be ContentType.objects.get(model='scanreportvalue').id,
                        # but that's not available from an Azure Function.
                        "content_type": 17,
                        "creation_type": "V",
                    }
                )

    # --------------------------------------------------------------------------------
    # Chunk the SRConcept data ready for upload, and then upload via the endpoint.
    logger.info(f"POST {len(concept_id_data)} concepts to table {table_name}")

    chunked_concept_id_data = helpers.perform_chunking(concept_id_data)
    logger.debug(f"chunked concepts list len: {len(chunked_concept_id_data)}")

    await post_chunks(
        chunked_concept_id_data,
        "scanreportconcepts",
        "concept",
        table_name=table_name,
        scan_report_id=scan_report_id,
    )

    logger.info("POST concepts all finished")

    reuse_existing_field_concepts(fieldnames_to_ids_dict, 15)
    reuse_existing_value_concepts(values_response_content, 17)


async def handle_single_table(
    current_table_name: str,
    current_table_id: str,
    scan_report_id: str,
    wb: Workbook,
    vocab_dictionary: Dict[str, Any],
):
    # This is the scenario where the line is empty, so we're at the end of
    # the table. Don't add a field entry, but process all those so far.

    # fields_response_content = post_field_entries(field_entries_to_post, scan_report_id)
    # TODO: Need to get this data from the API instead...
    # Create a dictionary with field names and field ids from the response
    # as key value pairs
    # e.g ("Field Name": Field ID)
    fieldnames_to_ids_dict = {
        str(element.get("name", None)): str(element.get("id", None))
        for element in fields_response_content
    }
    fieldids_to_names_dict = {
        str(element.get("id", None)): str(element.get("name", None))
        for element in fields_response_content
    }

    if current_table_name not in wb.sheetnames:
        helpers.process_failure(scan_report_id)
        raise ValueError(
            f"Attempting to access sheet '{current_table_name}'"
            f" in scan report, but no such sheet exists."
        )

    await process_values_for_table(
        vocab_dictionary,
        current_table_name,
        current_table_id,
        fieldnames_to_ids_dict,
        fieldids_to_names_dict,
        scan_report_id,
    )


async def process_all_fields_and_values(
    worksheet: Worksheet,
    wb: Workbook,
    table_name_to_id_map: Dict[str, str],
    vocab_dictionary: Dict[str, Any],
    scan_report_id: str,
):
    """
    Loop over all rows in Field Overview sheet.
    This is the same as looping over all fields in all tables.
    When the end of one table is reached, then post all the ScanReportFields
    and ScanReportValues associated to that table, then continue down the
    list of fields in tables.

    Args:

    """
    field_entries_to_post = []

    previous_row_value = None
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row + 2):
        # Guard against unnecessary rows beyond the last true row with contents
        if (previous_row_value is None or previous_row_value == "") and (
            row[0].value is None or row[0].value == ""
        ):
            break
        previous_row_value = row[0].value

        # If the row is not empty, then it is a field in a table, and should be added to
        # the list ready for processing at the end of this table.
        if row[0].value != "" and row[0].value is not None:
            current_table_name = row[0].value
            # Create ScanReportField entry
            field_entry = {
                "scan_report_table": table_name_to_id_map[current_table_name],
                "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "name": str(row[1].value),
                "description_column": str(row[2].value),
                "type_column": str(row[3].value),
                "max_length": row[4].value,
                "nrows": row[5].value,
                "nrows_checked": row[6].value,
                "fraction_empty": round(helpers.default_zero(row[7].value), 2),
                "nunique_values": row[8].value,
                "fraction_unique": round(helpers.default_zero(row[9].value), 2),
                "ignore_column": None,
            }
            # Append each entry to a list
            field_entries_to_post.append(field_entry)

        else:
            # This is the scenario where the line is empty, so we're at the end of
            # the table. Don't add a field entry, but process all those so far.
            # print("scan_report_field_entries >>>", field_entries_to_post)
            await handle_single_table(
                current_table_name,
                table_name_to_id_map[current_table_name],
                scan_report_id,
                wb,
                vocab_dictionary,
            )
            field_entries_to_post = []
    # Catch the final table if it wasn't already posted in the loop above - sometimes the iter_rows() seems to now allow you to go beyond the last row.
    if field_entries_to_post:
        await handle_single_table(
            current_table_name,
            table_name_to_id_map[current_table_name],
            scan_report_id,
            wb,
            vocab_dictionary,
        )


def main(msg: func.QueueMessage):
    """
    Processes a queue message.
    Unwraps the message content
    Gets the worbook, data_dictionary, and vocab_dictionary
    Creates concepts...
    """
    scan_report_blob, data_dictionary_blob, scan_report_id = helpers.unwrap_message(msg)

    wb, _, vocab_dictionary = blob_parser.parse_blobs(
        scan_report_blob, data_dictionary_blob
    )
    # Get the first sheet 'Field Overview',
    fo_ws = wb.worksheets[0]

    update_scan_report_status(scan_report_id, ScanReportStatus.PENDING)

    asyncio.run(
        process_all_fields_and_values(
            fo_ws,
            wb,
            table_name_to_id_map,
            vocab_dictionary,
            scan_report_id,
        )
    )

    update_scan_report_status(scan_report_id, ScanReportStatus.COMPLETE)

    wb.close()
    logger.info("Workbook successfully closed")
    return
