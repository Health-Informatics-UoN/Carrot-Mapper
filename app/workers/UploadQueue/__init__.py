import asyncio
import os
from collections import defaultdict
from typing import Any, Dict, List, Tuple

import azure.functions as func
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from shared.mapping.models import (
    ScanReportField,
    ScanReportTable,
    ScanReportValue,
)
from shared_code import blob_parser, helpers
from shared_code.db import (
    update_scan_report_status,
    UploadStatusConstant,
)
from shared_code.logger import logger

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "shared_code.django_settings")
import django

django.setup()


def _get_unique_table_names(worksheet: Worksheet) -> List[str]:
    """
    Extracts unique table names from the Field Overview worksheet.

    Args:
        worksheet: The worksheet containing table names.

    Returns:
        List[str]: A list of unique table names.
    """
    # Get all the table names in the order they appear in the Field Overview page
    table_names = []
    # Iterate over cells in the first column, but because we're in ReadOnly mode we
    # can't do that in the simplest manner.
    worksheet.reset_dimensions()
    worksheet.calculate_dimension(force=True)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        cell_value = row[0].value
        if cell_value and cell_value not in table_names:
            table_names.append(cell_value)
    return table_names


def _create_table_entry(table_name: str, id: str) -> ScanReportTable:
    """
    Creates a ScanReportTable entry.

    Args:
        table_name (str): The name of the table.
        id (str): The ID of the scan report.

    Returns:
        ScanReportTable: A model of the new ScanReporTable.
    """
    # Truncate table names because sheet names are truncated to 31 characters in Excel
    short_table_name = table_name[:31]

    return ScanReportTable(name=short_table_name, scan_report_id=id)


def _create_field_entry(row: Tuple[Cell], scan_report_table_id: str) -> ScanReportField:
    """
    Creates a ScanReportFieldEntry.

    Args:
        row (Tuple[Cell]): Row of data.
        scan_report_table_id (str): The ID of the scan report table

    Returns:
        ScanReportField: A dictionary representing the field entry.
    """
    return ScanReportField(
        scan_report_table_id=scan_report_table_id,
        name=str(row[1].value),
        description_column=str(row[2].value),
        type_column=str(row[3].value),
        max_length=row[4].value,
        nrows=row[5].value,
        nrows_checked=row[6].value,
        fraction_empty=round(helpers.default_zero(row[7].value), 2),
        nunique_values=row[8].value,
        fraction_unique=round(helpers.default_zero(row[9].value), 2),
    )


def _transform_scan_report_sheet_table(sheet: Worksheet) -> defaultdict[Any, List]:
    """
    Transforms a worksheet data into a JSON like format.

    Args:
        sheet (Worksheet): Sheet of data to transform

    Returns:
        defaultdict[Any, List]: The transformed data.
    """
    logger.debug("Start process_scan_report_sheet_table")

    sheet.reset_dimensions()
    sheet.calculate_dimension(force=True)
    # Get header entries (skipping every second column which is just 'Frequency')
    # So sheet_headers = ['a', 'b']
    first_row = sheet[1]
    sheet_headers = [cell.value for cell in first_row[::2]]

    # Set up an empty defaultdict, and fill it with one entry per header (i.e. one
    # per column)
    # Append each entry's value with the tuple (value, frequency) so that we end up
    # with each entry containing one tuple per non-empty entry in the column.
    #
    # This will give us
    #
    # ordereddict({'a': [('apple', 20), ('banana', 3), ('pear', 12)],
    #              'b': [('orange', 5), ('plantain', 50)]})

    d = defaultdict(list)
    # Iterate over all rows beyond the header - use the number of sheet_headers*2 to
    # set the maximum column rather than relying on sheet.max_col as this is not
    # always reliably updated by Excel etc.
    for row in sheet.iter_rows(
        min_col=1,
        max_col=len(sheet_headers) * 2,
        min_row=2,
        max_row=sheet.max_row,
        values_only=True,
    ):
        # Set boolean to track whether we hit a blank row for early exit below.
        this_row_empty = True
        # Iterate across the pairs of cells in the row. If the pair is non-empty,
        # then add it to the relevant dict entry.
        for header, cell, freq in zip(sheet_headers, row[::2], row[1::2]):
            if (cell != "" and cell is not None) or (freq != "" and freq is not None):
                d[header].append((str(cell), freq))
                this_row_empty = False
        # This will trigger if we hit a row that is entirely empty. Short-circuit
        # to exit early here - this saves us from situations where sheet.max_row is
        # incorrectly set (too large)
        if this_row_empty:
            break

    logger.debug("Finish process_scan_report_sheet_table")
    return d


def _create_value_entries(
    values_details: List[Dict[str, Any]], fields: list[ScanReportField]
) -> List[ScanReportValue]:
    """
    Create value entries based on values_details and fieldnames_to_ids_dict.

    Args:
        values_details (List[Dict[str, Any]]): A list of dictionaries of the value
            details for each fieldname-value pair.
        fields (List[ScanReportField]): A list of SCan Report Fields.

    Returns:
        List[ScanReportValue]: A list of ScanReportValues.
    """
    return [
        ScanReportValue(
            value=entry["full_value"][:127],
            frequency=int(entry["frequency"]),
            value_description=entry["val_desc"],
            scan_report_field=next(f for f in fields if f.name == entry["fieldname"]),
        )
        for entry in values_details
    ]


def _apply_data_dictionary(
    values_details: List[Dict[str, Any]], data_dictionary: Dict[Any, Dict]
) -> None:
    """
    Apply data dictionary to update value descriptions in values_details from the data dict.

    Args:
        values_details (List[Dict[str, Any]]): A list of dictionaries with the value
            details for each fieldname-value pair.
        data_dictionary (Dict[Any, Dict]): A dictionary mapping table names to dictionaries
            containing fieldname-value mappings and their corresponding value descriptions.

    Returns:
        None
    """
    for entry in values_details:
        table_data = data_dictionary.get(str(entry["table"]))
        if table_data and table_data.get(str(entry["fieldname"])):
            entry["val_desc"] = table_data[str(entry["fieldname"])].get(
                str(entry["full_value"])
            )


def _assign_order(values_details: List[Dict[str, Any]]) -> None:
    """
    Add "order" field to each entry to enable correctly-ordered recombination at the end.

    Args:
        values_details (List[Dict[str, Any]]): List of value details to transform.

    Returns:
        None
    """
    for entry_number, entry in enumerate(values_details):
        entry["order"] = entry_number


def _create_values_details(
    fieldname_value_freq: Dict[str, Tuple[str]],
    table_name: str,
) -> List[Dict[str, Any]]:
    """
    Create value details for each fieldname-value pair.

    Args:
        fieldname_value_freq (Dict[str, Tuple[str]]): A dictionary mapping fieldnames to
            tuples of value-frequency pairs.
        table_name (str): The name of the table.

    Returns:
        List[Dict[str, Any]]: A list of dictionaries containing the value details for each
            fieldname-value pair.
    """
    values_details = []
    for fieldname, value_freq_tuples in fieldname_value_freq.items():
        for full_value, frequency in value_freq_tuples:
            try:
                frequency = int(frequency)
            except (ValueError, TypeError):
                frequency = 0
            values_details.append(
                {
                    "full_value": full_value,
                    "frequency": frequency,
                    "fieldname": fieldname,
                    "table": table_name,
                    "val_desc": None,
                }
            )
    return values_details


async def _add_SRValues_and_value_descriptions(
    fieldname_value_freq_dict: Dict[str, Tuple[str]],
    current_table_name: str,
    data_dictionary: Dict[Any, Dict],
    fields: list[ScanReportField],
) -> None:
    """
    Add ScanReportValues and value descriptions to the values_details list.

    Args:
        fieldname_value_freq_dict: A dictionary containing field names as keys and
            value-frequency tuples as values.
        current_table_name: The name of the current table.
        data_dictionary: The data dictionary containing field-value descriptions.
        fields: A list of Scan Report Fields.

    Returns:
        The response content after posting the values.
    """
    values_details = _create_values_details(
        fieldname_value_freq_dict, current_table_name
    )
    logger.debug("Assign order")
    _assign_order(values_details)

    # --------------------------------------------------------------------------------
    # Update val_desc of each SRField entry if it has a value description from the
    # data dictionary
    if data_dictionary:
        logger.debug("apply data dictionary")
        _apply_data_dictionary(values_details, data_dictionary)

    # Convert basic information about SRValues into entries
    logger.debug("create value_entries_to_post")
    value_entries = _create_value_entries(values_details, fields)
    await ScanReportValue.objects.abulk_create(value_entries)


async def _handle_single_table(
    current_table_name: str,
    field_entries: list[ScanReportField],
    scan_report_id: str,
    workbook: Workbook,
    data_dictionary: Dict[Any, Dict],
) -> None:
    """
    Handle creating a single table values.

    Args:
        field_entries (List[Dict[str, str]]): List of field entries to create.
        scan_report_id (str): ID of the scan report to attach to.

    Raises:
        Exception: ValueError: Trying to access a sheet in the workbook that does not exist.
    """
    fields = await ScanReportField.objects.abulk_create(field_entries)

    if current_table_name not in workbook.sheetnames:
        update_scan_report_status(scan_report_id, UploadStatusConstant.FAILED)
        raise ValueError(
            f"Attempting to access sheet '{current_table_name}'"
            f" in scan report, but no such sheet exists."
        )

    # Go to Table sheet to process all the values from the sheet
    sheet = workbook[current_table_name]

    fieldname_value_freq_dict = _transform_scan_report_sheet_table(sheet)
    await _add_SRValues_and_value_descriptions(
        fieldname_value_freq_dict,
        current_table_name,
        data_dictionary,
        fields,
    )


def _create_tables(worksheet: Worksheet, id: str) -> list[ScanReportTable]:
    """
    Creates tables extracted from the Field Overview worksheet.

    For each table name create a ScanReportTable.

    Args:
        worksheet (Worksheet): The worksheet containing table names.
        id (str): The ID of the scan report.

    Returns:
        list[ScanReportTable]: A list of the ScanReportTables created.
    """
    table_names = _get_unique_table_names(worksheet)
    logger.info(f"TABLES NAMES >>> {table_names}")
    table_models = [_create_table_entry(name, id) for name in table_names]
    return ScanReportTable.objects.bulk_create(table_models)


async def _create_fields(
    worksheet: Worksheet,
    workbook: Workbook,
    id: str,
    tables: list[ScanReportTable],
    data_dictionary: Dict[Any, Dict],
) -> None:
    """
    Creates fields extracted from the Field Overview worksheet.

    Loop over all rows in Field Overview sheet.
    This is the same as looping over all fields in all tables.
    When the end of one table is reached, then post all the ScanReportFields
    and ScanReportValues associated to that table, then continue down the
    list of fields in tables.

    Args:
        worksheet (Worksheet): The worksheet containing table names.
        id (str): Scan Report ID to POST to
    """
    field_entries_to_post = []

    previous_row_value = None
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row + 2):
        # Guard against unnecessary rows beyond the last true row with contents
        if (previous_row_value is None or previous_row_value == "") and (
            row[0].value is None or row[0].value == ""
        ):
            break
        previous_row_value = row[0].value

        # If the row is not empty, then it is a field in a table, and should be added to
        # the list ready for processing at the end of this table.
        if row[0].value != "" and row[0].value is not None:
            current_table_name = row[0].value
            table = next(t for t in tables if t.name == current_table_name)
            # get the current table in the list of tables by name.

            field_entry = _create_field_entry(row, table.pk)
            field_entries_to_post.append(field_entry)
        else:
            # This is the scenario where the line is empty, so we're at the end of
            # the table. Don't add a field entry, but process all those so far.
            await _handle_single_table(
                str(current_table_name),
                field_entries_to_post,
                id,
                workbook,
                data_dictionary,
            )
            field_entries_to_post = []

    # Catch the final table if it wasn't already posted in the loop above -
    # sometimes the iter_rows() seems to now allow you to go beyond the last row.
    if field_entries_to_post:
        await _handle_single_table(
            str(current_table_name),
            field_entries_to_post,
            id,
            workbook,
            data_dictionary,
        )


def _handle_failure(msg: func.QueueMessage, scan_report_id: str) -> None:
    """
    Handles failure scenarios where the message has been dequeued more than once.

    Args:
        msg (func.QueueMessage): The message received from the queue.
        scan_report_id (str): The ID of the scan report.

    Raises:
        ValueError: If the dequeue count of the message exceeds 1.
    """
    logger.info(f"dequeue_count {msg.dequeue_count}")

    if msg.dequeue_count == 2:
        update_scan_report_status(scan_report_id, UploadStatusConstant.FAILED)
    if msg.dequeue_count > 1:
        raise ValueError("dequeue_count > 1")


def main(msg: func.QueueMessage) -> None:
    """
    Processes a queue message
    Unwraps the message content
    Gets the workbook and data dictionary.
    Creates the scan report tables.
    Creates the scan report fields.
    Updates the scan report status accordingly.

    Args:
        msg (func.QueueMessage): The message received from the queue.
    """
    scan_report_blob, data_dictionary_blob, scan_report_id, _ = helpers.unwrap_message(
        msg
    )
    _handle_failure(msg, scan_report_id)

    update_scan_report_status(scan_report_id, UploadStatusConstant.IN_PROGRESS)

    wb = blob_parser.get_scan_report(scan_report_blob)
    data_dictionary, _ = blob_parser.get_data_dictionary(data_dictionary_blob)

    # Get the first sheet 'Field Overview',
    # to populate ScanReportTable & ScanReportField models
    fo_ws = wb.worksheets[0]

    table_name_to_id_map = _create_tables(fo_ws, scan_report_id)
    asyncio.run(
        _create_fields(fo_ws, wb, scan_report_id, table_name_to_id_map, data_dictionary)
    )

    update_scan_report_status(scan_report_id, UploadStatusConstant.COMPLETE)
