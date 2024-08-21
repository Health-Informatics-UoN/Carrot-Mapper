import csv
from collections import Counter
from io import BytesIO, StringIO

import openpyxl
from django import forms
from django.contrib.auth import password_validation
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.forms.models import ModelChoiceField
from shared.mapping.models import (
    Dataset,
    ScanReport,
    ScanReportField,
    VisibilityChoices,
)


class ShowNameChoiceField(ModelChoiceField):
    def label_from_instance(self, obj):
        return obj.name


class ScanReportForm(forms.Form):
    dataset = forms.CharField(
        label="Scan Report name",
        widget=forms.TextInput(attrs={"class": "form-control"}),
    )
    scan_report_file = forms.FileField(
        label="WhiteRabbit ScanReport",
        widget=forms.FileInput(attrs={"class": "form-control"}),
    )

    data_dictionary_file = forms.FileField(
        label="Data Dictionary",
        widget=forms.FileInput(attrs={"class": "form-control"}),
        required=False,
    )
    parent_dataset = ShowNameChoiceField(
        label="Dataset",
        queryset=Dataset.objects.order_by("name"),
        widget=forms.Select(attrs={"class": "form-control"}),
    )
    viewers = forms.ModelMultipleChoiceField(
        label="Viewers",
        queryset=User.objects.order_by("username"),
        widget=forms.SelectMultiple(attrs={"class": "form-control"}),
        required=False,
    )
    editors = forms.ModelMultipleChoiceField(
        label="Editors",
        queryset=User.objects.order_by("username"),
        widget=forms.SelectMultiple(attrs={"class": "form-control"}),
        required=False,
    )
    visibility = forms.ChoiceField(
        label="Visibility",
        choices=VisibilityChoices.choices,
        widget=forms.Select(attrs={"class": "form-control"}),
    )

    class Meta:
        model = ScanReport
        fields = (
            "dataset",
            "scan_report_file",
            "parent_dataset",
            "editors",
            "viewers",
            "visibility",
        )

    def clean_data_dictionary_file(self):
        data_dictionary = self.cleaned_data.get("data_dictionary_file")

        if data_dictionary is None:
            return data_dictionary

        if not str(data_dictionary).endswith(".csv"):
            raise ValidationError(
                "You have attempted to upload a data dictionary "
                "which is not in CSV format. "
                "Please upload a .csv file."
            )

        csv_reader = csv.reader(StringIO(data_dictionary.read().decode("utf-8-sig")))

        errors = []

        # Check first line for correct headers to columns
        header_line = next(csv_reader)
        if header_line != ["csv_file_name", "field_name", "code", "value"]:
            errors.append(
                ValidationError(
                    f"Dictionary file has incorrect first line. "
                    f"It must be ['csv_file_name', "
                    f"'field_name', 'code', 'value'], but you "
                    f"supplied {header_line}. If this error is "
                    f"showing extra '' elements, this indicates "
                    f"that another line has >4 elements, "
                    f"which will need to be corrected."
                )
            )

        # Check all rows have either 3 or 4 non-empty elements, and only the 4th can be empty.
        # Start from 2 because we want to use 1-indexing _and_ skip the first row which was
        # processed above.
        for line_no, line in enumerate(csv_reader, start=2):
            line_length_nonempty = len([element for element in line if element != ""])
            if line_length_nonempty not in [3, 4]:
                errors.append(
                    ValidationError(
                        f"Dictionary has "
                        f"{line_length_nonempty} "
                        f"values in line {line_no} ({line}). "
                        f"All lines must "
                        f"have either 3 or 4 entries."
                    )
                )
            # Check for whether any of the first 3 elements are empty
            for element_no, element in enumerate(line[:3], start=1):
                if element == "":
                    errors.append(
                        ValidationError(
                            f"Dictionary has an empty element "
                            f"in column {element_no} in line "
                            f"{line_no}. "
                            f"Only the 4th element in any line "
                            f"may be empty."
                        )
                    )

        if errors:
            raise ValidationError(errors)

        return data_dictionary

    def run_fast_consistency_checks(self, wb):
        """
        Runs a number of consistency checks on the provided workbook. The aim is to
        return quickly if there is an issue with the data, and provide feedback to the
        user so they can fix the issue.
        """
        errors = []
        # Get the first sheet 'Field Overview'
        fo_ws = wb.worksheets[0]

        # Grab the scan report columns from the first worksheet
        # Define what the column headings should be
        source_headers = [header.value for header in fo_ws[1]]

        expected_headers = [
            "Table",
            "Field",
            "Description",
            "Type",
            "Max length",
            "N rows",
            "N rows checked",
            "Fraction empty",
            "N unique values",
            "Fraction unique",
        ]

        # Check if source headers match the expected headers. Allow unexpected
        # headers after these. This means old Scan Reports with Flag and Classification
        # columns will be handled cleanly.
        if not source_headers[:10] == expected_headers:
            errors.append(
                ValidationError(
                    f"Please check the following columns exist "
                    f"in the Scan Report (Field Overview sheet) "
                    f"in this order: "
                    f"Table, Field, Description, Type, "
                    f"Max length, N rows, N rows checked, "
                    f"Fraction empty, N unique values, "
                    f"Fraction unique. "
                    f"You provided \n{source_headers[:10]}"
                )
            )
            raise ValidationError(errors)

        # Check tables are correctly separated in FO - a single empty line between each
        # table
        cell_above = fo_ws["A"][1]
        for cell in fo_ws["A"][1:]:
            if (
                cell.value != cell_above.value
                and (cell.value != "" and cell.value is not None)
                and (cell_above.value != "" and cell_above.value is not None)
            ) or (cell.value == "" and cell_above.value == ""):
                errors.append(
                    ValidationError(
                        f"At {cell}, tables in Field Overview "
                        f"table are not correctly separated by "
                        f"a single line. "
                        f"Note: There should be no separator "
                        f"line between the header row and the "
                        f"first row of the first table."
                    )
                )
            cell_above = cell

        if errors:
            raise ValidationError(errors)

        # Now that we're happy that the FO sheet is correctly formatted, we can move
        # on to comparing its contents to the sheets

        # Check tables in FO match supplied sheets
        table_names = set(
            cell.value
            for cell in fo_ws["A"][1:]
            if (cell.value != "" and cell.value is not None)
        )
        # Drop "Table Overview" and "_" sheetnames if present, as these are never used.
        table_names.difference_update(["Table Overview", "_"])

        # "Field Overview" is the only required sheet that is not a table name.
        expected_sheetnames = list(table_names) + ["Field Overview"]

        # Get names of sheet from workbook
        actual_sheetnames = set(wb.sheetnames)
        # Drop "Table Overview" and "_" sheetnames if present, as these are never used.
        actual_sheetnames.difference_update(["Table Overview", "_"])

        if sorted(actual_sheetnames) != sorted(expected_sheetnames):
            sheets_only = set(actual_sheetnames).difference(expected_sheetnames)
            fo_only = set(expected_sheetnames).difference(actual_sheetnames)
            errors.append(
                ValidationError(
                    "Tables in Field Overview sheet do not "
                    "match the sheets supplied."
                )
            )
            if sheets_only:
                errors.append(
                    ValidationError(
                        f"{sheets_only} are sheets that do not "
                        f"have matching entries in first column "
                        f"of the Field Overview sheet. "
                    )
                )
            if fo_only:
                errors.append(
                    ValidationError(
                        f"{fo_only} are table names in first "
                        f"column of Field Overview sheet but do "
                        f"not have matching sheets supplied."
                    )
                )

        if errors:
            raise ValidationError(errors)

        # Loop over the rows, and for each table, once we reach the end of the table,
        # compare the fields provided with the fields in the associated sheet
        current_table_fields = []
        last_value = None
        for row in fo_ws.iter_rows(min_row=2):
            # Loop over rows, collecting all fields in each table in turn
            if row[0].value == "" or row[0].value is None:
                # We're at the end of the table, so process
                # Firstly, check that we're not two empty lines in a row - if so,
                # then we're beyond the last true value and iter_rows is just giving
                # us spurious rows. Abort early.
                if last_value == "" or last_value is None:
                    break
                # Get all field names from the associated sheet, by grabbing the first
                # row, and then grabbing every second column value (because the
                # alternate columns should be 'Frequency'
                table_sheet_fields = [
                    cell.value for cell in next(wb[current_table_name].rows)
                ][::2]

                # Check for multiple columns in a single sheet with the same name
                count_table_sheet_fields = Counter(table_sheet_fields)
                for field in count_table_sheet_fields:
                    if count_table_sheet_fields[field] > 1:
                        errors.append(
                            ValidationError(
                                f"Sheet '{current_table_name}' "
                                f"contains more than one field "
                                f"with the name '{field}'. "
                                f"Field names must be unique "
                                f"within a table."
                            )
                        )

                # Check for multiple fields with the same name associated to a single
                # table in the Field Overview sheet
                count_current_table_fields = Counter(current_table_fields)
                for field in count_current_table_fields:
                    if count_current_table_fields[field] > 1:
                        errors.append(
                            ValidationError(
                                f"Field Overview sheet contains "
                                f"more than one field with the "
                                f"name '{field}' against the "
                                f"table '{current_table_name}'. "
                                f"Field names must be unique "
                                f"within a table."
                            )
                        )

                # Check for any fields that are in only one of the Field Overview and
                # the associated sheet
                if sorted(table_sheet_fields) != sorted(current_table_fields):
                    sheet_only = set(table_sheet_fields).difference(
                        current_table_fields
                    )
                    fo_only = set(current_table_fields).difference(table_sheet_fields)
                    errors.append(
                        ValidationError(
                            f"Fields in Field Overview against "
                            f"table {current_table_name} do not "
                            f"match fields in the associated "
                            f"sheet. "
                        )
                    )
                    if sheet_only:
                        errors.append(
                            ValidationError(
                                f"{sheet_only} exist in the "
                                f"'{current_table_name}' sheet "
                                f"but there are no matching "
                                f"entries in the second column "
                                f"of the Field Overview sheet "
                                f"in the rows associated to the "
                                f"table '{current_table_name}'. "
                                f""
                            )
                        )
                    if fo_only:
                        errors.append(
                            ValidationError(
                                f"{fo_only} exist in second "
                                f"column of Field Over"
                                f"view sheet against the table "
                                f"'{current_table_name}' but "
                                f"there are no matching column "
                                f"names in the associated sheet "
                                f"'{current_table_name}'."
                            )
                        )

                # Reset the list of fields associated to this table as we iterate down
                # the FO sheet.
                current_table_fields = []
            else:
                # Update current list of field names and the current table name - we can
                # trust the table name not to change in this case due to the earlier
                # check for empty lines between tables in the FO sheet.
                current_table_fields.append(row[1].value)
                current_table_name = row[0].value

            last_value = row[0].value

        if errors:
            raise ValidationError(errors)

        return True

    def clean_scan_report_file(self):
        scan_report = self.cleaned_data.get("scan_report_file")

        if not str(scan_report).endswith(".xlsx"):
            raise ValidationError(
                "You have attempted to upload a scan report which "
                "is not in XLSX format. Please upload a .xlsx file."
            )

        # Load in the Excel sheet, grab the first workbook
        file_in_memory = scan_report.read()
        wb = openpyxl.load_workbook(filename=BytesIO(file_in_memory), data_only=True)

        self.run_fast_consistency_checks(wb)

        return scan_report


class UserCreateForm(UserCreationForm):
    email = forms.EmailField(
        required=True, label="Email", error_messages={"exists": "Oops"}
    )

    class Meta:
        model = User
        fields = ("username", "email", "password1", "password2")

    def save(self, commit=True):
        user = super(UserCreateForm, self).save(commit=False)
        user.email = self.cleaned_data["email"]
        if commit:
            user.save()
        return user

    def clean_email(self):
        if User.objects.filter(email=self.cleaned_data["email"]).exists():
            raise ValidationError(self.fields["email"].error_messages["exists"])
        return self.cleaned_data["email"]


class PasswordChangeForm(forms.Form):
    old_password = forms.CharField(
        label="Old Password",
        widget=forms.PasswordInput(
            attrs={"autocomplete": "current-password", "autofocus": True}
        ),
    )
    new_password1 = forms.CharField(
        label=("New password"),
        widget=forms.PasswordInput(attrs={"autocomplete": "new-password"}),
        validators=[password_validation.validate_password],
    )

    new_password2 = forms.CharField(
        label=("Confirm New password"),
        widget=forms.PasswordInput(attrs={"autocomplete": "new-password"}),
        validators=[password_validation.validate_password],
    )


class ScanReportAssertionForm(forms.Form):
    negative_assertion = forms.CharField(
        label="Negative Assertions",
        widget=forms.TextInput(attrs={"class": "form-control"}),
    )


class NLPForm(forms.Form):
    user_string = forms.CharField(label="")


class ScanReportFieldConceptForm(forms.Form):
    scan_report_field_id = forms.IntegerField()

    concept_id = forms.IntegerField()


class ScanReportValueConceptForm(forms.Form):
    scan_report_value_id = forms.IntegerField()

    concept_id = forms.IntegerField()


class ScanReportFieldForm(forms.ModelForm):
    class Meta:
        model = ScanReportField
        fields = (
            # "is_patient_id",
            "is_ignore",
            "pass_from_source",
            "description_column",
        )
        widgets = {
            "description_column": forms.TextInput(attrs={"class": "form-control"})
        }
