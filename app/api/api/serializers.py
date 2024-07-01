from django.contrib.auth.models import User
from drf_dynamic_fields import DynamicFieldsMixin
from mapping.permissions import has_editorship, is_admin, is_az_function_user
from mapping.services_rules import analyse_concepts, get_mapping_rules_json
from rest_framework import serializers
from rest_framework.exceptions import NotFound, PermissionDenied, ParseError
from shared.data.models import (
    ClassificationSystem,
    DataDictionary,
    DataPartner,
    Dataset,
    MappingRule,
    OmopField,
    OmopTable,
    Project,
    ScanReport,
    ScanReportConcept,
    ScanReportField,
    ScanReportTable,
    ScanReportValue,
)
from shared.data.omop import (
    Concept,
    ConceptAncestor,
    ConceptClass,
    ConceptRelationship,
    ConceptSynonym,
    Domain,
    DrugStrength,
    Vocabulary,
)
from openpyxl.workbook.workbook import Workbook
from shared.data.models import Dataset, ScanReport, ScanReportField, VisibilityChoices
import csv
from io import BytesIO, StringIO
import openpyxl
from collections import Counter


class DataPartnerSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = DataPartner
        fields = "__all__"


class ConceptSerializer(serializers.ModelSerializer):
    class Meta:
        model = Concept
        fields = "__all__"


class VocabularySerializer(serializers.ModelSerializer):
    class Meta:
        model = Vocabulary
        fields = "__all__"


class ConceptRelationshipSerializer(serializers.ModelSerializer):
    class Meta:
        model = ConceptRelationship
        fields = "__all__"


class ConceptAncestorSerializer(serializers.ModelSerializer):
    class Meta:
        model = ConceptAncestor
        fields = "__all__"


class ConceptClassSerializer(serializers.ModelSerializer):
    class Meta:
        model = ConceptClass
        fields = "__all__"


class ConceptSynonymSerializer(serializers.ModelSerializer):
    class Meta:
        model = ConceptSynonym
        fields = "__all__"


class DomainSerializer(serializers.ModelSerializer):
    class Meta:
        model = Domain
        fields = "__all__"


class DrugStrengthSerializer(serializers.ModelSerializer):
    class Meta:
        model = DrugStrength
        fields = "__all__"


class UserSerializer(serializers.ModelSerializer):
    class Meta:
        model = User
        fields = ("id", "username")


class ScanReportViewSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    def validate(self, data):
        if request := self.context.get("request"):
            if ds := data.get("parent_dataset"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(ds, request)
                    or has_editorship(ds, request)
                ):
                    raise PermissionDenied(
                        "You must be an admin of the parent dataset to add a new scan report to it.",
                    )
            else:
                raise NotFound("Could not find parent dataset.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report."
            )
        return super().validate(data)

    class Meta:
        model = ScanReport
        fields = "__all__"


class ScanReportViewSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):
    """
    Serializer for the ScanReportViewV2, for version 2.
    Args:
        self: The instance of the class.
        data: The data to be validated.
    Returns:
        dict: The validated data for the scan report.
    Raises:
        serializers.ValidationError: If the request context is missing.
        PermissionDenied: If the user does not have the required permissions.
        NotFound: If the parent dataset is not found.
    """

    parent_dataset = serializers.SerializerMethodField()
    data_partner = serializers.SerializerMethodField()

    class Meta:
        model = ScanReport
        fields = (
            "id",
            "name",
            "dataset",
            "parent_dataset",
            "data_partner",
            "status",
            "created_at",
            "hidden",
        )

    def get_parent_dataset(self, obj):
        return obj.parent_dataset.name if obj.parent_dataset else None

    def get_data_partner(self, obj):
        return (
            obj.parent_dataset.data_partner.name
            if obj.parent_dataset.data_partner
            else None
        )


class ScanReportFilesSerializer(DynamicFieldsMixin, serializers.ModelSerializer):

    scan_report_file = serializers.FileField(write_only=True)
    data_dictionary_file = serializers.FileField(
        write_only=True, required=False, allow_empty_file=True
    )

    class Meta:
        model = ScanReport
        fields = (
            "scan_report_file",
            "data_dictionary_file",
        )

    def validate_data_dictionary_file(self, value):
        data_dictionary = value

        if str(data_dictionary) == "undefined":
            return data_dictionary

        if not str(data_dictionary).endswith(".csv"):
            raise ParseError(
                "You have attempted to upload a data dictionary "
                "which is not in CSV format. "
                "Please upload a .csv file."
            )

        csv_reader = csv.reader(StringIO(data_dictionary.read().decode("utf-8-sig")))

        errors = []

        # Check first line for correct headers to columns
        header_line = next(csv_reader)
        if header_line != ["csv_file_name", "field_name", "code", "value"]:
            raise ParseError(
                f"Dictionary file has incorrect first line. "
                f"It must be ['csv_file_name', "
                f"'field_name', 'code', 'value'], but you "
                f"supplied {header_line}. If this error is "
                f"showing extra elements, this indicates "
                f"that another line has >4 elements, "
                f"which will need to be corrected."
            )

        # Check all rows have either 3 or 4 non-empty elements, and only the 4th can be empty.
        # Start from 2 because we want to use 1-indexing _and_ skip the first row which was
        # processed above.
        for line_no, line in enumerate(csv_reader, start=2):
            line_length_nonempty = len([element for element in line if element != ""])
            if line_length_nonempty not in [3, 4]:
                errors.append(
                    ParseError(
                        f"Dictionary has "
                        f"{line_length_nonempty} "
                        f"values in line {line_no} ({line}). "
                        f"All lines must "
                        f"have either 3 or 4 entries."
                    )
                )
            # Check for whether any of the first 3 elements are empty
            for element_no, element in enumerate(line[:3], start=1):
                if element == "":
                    errors.append(
                        ParseError(
                            f"Dictionary has an empty element "
                            f"in column {element_no} in line "
                            f"{line_no}. "
                            f"Only the 4th element in any line "
                            f"may be empty."
                        )
                    )

        if errors:
            raise ParseError(errors)

        return data_dictionary

    def run_fast_consistency_checks(self, wb: Workbook):
        """
        This function performs a series of consistency checks on the provided Excel workbook.
        The checks are designed to quickly identify and provide feedback on common data issues,
        enabling the user to correct them.

        If any of these checks fail, one or a list of ParseError will be raised with a message detailing the issue.

        Args:
            wb (Workbook): The Excel workbook to check.

        Returns:
            True if all checks pass.

        Raises:
            ParseError: Validation checks have failed.
        """
        errors = []
        # Get the first sheet 'Field Overview'
        fo_ws = wb.worksheets[0]

        # Grab the scan report columns from the first worksheet
        # Define what the column headings should be
        source_headers = [header.value for header in fo_ws[1]]

        expected_headers = [
            "Table",
            "Field",
            "Description",
            "Type",
            "Max length",
            "N rows",
            "N rows checked",
            "Fraction empty",
            "N unique values",
            "Fraction unique",
        ]

        # Check if source headers match the expected headers. Allow unexpected
        # headers after these. This means old Scan Reports with Flag and Classification
        # columns will be handled cleanly.
        if not source_headers[:10] == expected_headers:
            errors.append(
                ParseError(
                    f"Please check the following columns exist "
                    f"in the Scan Report (Field Overview sheet) "
                    f"in this order: "
                    f"Table, Field, Description, Type, "
                    f"Max length, N rows, N rows checked, "
                    f"Fraction empty, N unique values, "
                    f"Fraction unique. "
                    f"You provided \n{source_headers[:10]}"
                )
            )
            raise ParseError(errors)

        # Check tables are correctly separated in FO - a single empty line between each
        # table
        cell_above = fo_ws["A"][1]
        for cell in fo_ws["A"][1:]:
            if (
                cell.value != cell_above.value
                and (cell.value != "" and cell.value is not None)
                and (cell_above.value != "" and cell_above.value is not None)
            ) or (cell.value == "" and cell_above.value == ""):
                errors.append(
                    ParseError(
                        f"At the cell with value {cell.value}, tables in Field Overview "
                        f"table are not correctly separated by "
                        f"a single line. "
                        f"Note: There should be no separator "
                        f"line between the header row and the "
                        f"first row of the first table."
                    )
                )
            cell_above = cell

        if errors:
            raise ParseError(errors)

        # Now that we're happy that the FO sheet is correctly formatted, we can move
        # on to comparing its contents to the sheets

        # Check tables in FO match supplied sheets
        table_names = set(
            cell.value
            for cell in fo_ws["A"][1:]
            if (cell.value != "" and cell.value is not None)
        )
        # Drop "Table Overview" and "_" sheetnames if present, as these are never used.
        table_names.difference_update(["Table Overview", "_"])

        # "Field Overview" is the only required sheet that is not a table name.
        expected_sheetnames = list(table_names) + ["Field Overview"]

        # Get names of sheet from workbook
        actual_sheetnames = set(wb.sheetnames)
        # Drop "Table Overview" and "_" sheetnames if present, as these are never used.
        actual_sheetnames.difference_update(["Table Overview", "_"])

        if sorted(actual_sheetnames) != sorted(expected_sheetnames):
            sheets_only = set(actual_sheetnames).difference(expected_sheetnames)
            fo_only = set(expected_sheetnames).difference(actual_sheetnames)
            errors.append(
                ParseError(
                    "Tables in Field Overview sheet do not "
                    "match the sheets supplied."
                )
            )
            if sheets_only:
                errors.append(
                    ParseError(
                        f"{sheets_only} are sheets that do not "
                        f"have matching entries in first column "
                        f"of the Field Overview sheet. "
                    )
                )
            if fo_only:
                errors.append(
                    ParseError(
                        f"{fo_only} are table names in first "
                        f"column of Field Overview sheet but do "
                        f"not have matching sheets supplied."
                    )
                )

        if errors:
            raise ParseError(errors)

        # Loop over the rows, and for each table, once we reach the end of the table,
        # compare the fields provided with the fields in the associated sheet
        current_table_fields = []
        last_value = None
        for row in fo_ws.iter_rows(min_row=2):
            # Loop over rows, collecting all fields in each table in turn
            if row[0].value == "" or row[0].value is None:
                # We're at the end of the table, so process
                # Firstly, check that we're not two empty lines in a row - if so,
                # then we're beyond the last true value and iter_rows is just giving
                # us spurious rows. Abort early.
                if last_value == "" or last_value is None:
                    break
                # Get all field names from the associated sheet, by grabbing the first
                # row, and then grabbing every second column value (because the
                # alternate columns should be 'Frequency'
                table_sheet_fields = [
                    cell.value for cell in next(wb[current_table_name].rows)
                ][::2]

                # Check for multiple columns in a single sheet with the same name
                count_table_sheet_fields = Counter(table_sheet_fields)
                for field in count_table_sheet_fields:
                    if count_table_sheet_fields[field] > 1:
                        errors.append(
                            ParseError(
                                f"Sheet '{current_table_name}' "
                                f"contains more than one field "
                                f"with the name '{field}'. "
                                f"Field names must be unique "
                                f"within a table."
                            )
                        )

                # Check for multiple fields with the same name associated to a single
                # table in the Field Overview sheet
                count_current_table_fields = Counter(current_table_fields)
                for field in count_current_table_fields:
                    if count_current_table_fields[field] > 1:
                        errors.append(
                            ParseError(
                                f"Field Overview sheet contains "
                                f"more than one field with the "
                                f"name '{field}' against the "
                                f"table '{current_table_name}'. "
                                f"Field names must be unique "
                                f"within a table."
                            )
                        )

                # Check for any fields that are in only one of the Field Overview and
                # the associated sheet
                if sorted(table_sheet_fields) != sorted(current_table_fields):
                    sheet_only = set(table_sheet_fields).difference(
                        current_table_fields
                    )
                    fo_only = set(current_table_fields).difference(table_sheet_fields)
                    errors.append(
                        ParseError(
                            f"Fields in Field Overview against "
                            f"table {current_table_name} do not "
                            f"match fields in the associated "
                            f"sheet. "
                        )
                    )
                    if sheet_only:
                        errors.append(
                            ParseError(
                                f"{sheet_only} exist in the "
                                f"'{current_table_name}' sheet "
                                f"but there are no matching "
                                f"entries in the second column "
                                f"of the Field Overview sheet "
                                f"in the rows associated to the "
                                f"table '{current_table_name}'. "
                                f""
                            )
                        )
                    if fo_only:
                        errors.append(
                            ParseError(
                                f"{fo_only} exist in second "
                                f"column of Field Over"
                                f"view sheet against the table "
                                f"'{current_table_name}' but "
                                f"there are no matching column "
                                f"names in the associated sheet "
                                f"'{current_table_name}'."
                            )
                        )

                # Reset the list of fields associated to this table as we iterate down
                # the FO sheet.
                current_table_fields = []
            else:
                # Update current list of field names and the current table name - we can
                # trust the table name not to change in this case due to the earlier
                # check for empty lines between tables in the FO sheet.
                current_table_fields.append(row[1].value)
                current_table_name = row[0].value

            last_value = row[0].value

        if errors:
            raise ParseError(errors)

        return True

    def validate_scan_report_file(self, value):
        scan_report = value

        if not str(scan_report).endswith(".xlsx"):
            raise ParseError(
                "You have attempted to upload a scan report which "
                "is not in XLSX format. Please upload a .xlsx file."
            )

        # Load in the Excel sheet, grab the first workbook
        file_in_memory = scan_report.read()
        wb = openpyxl.load_workbook(filename=BytesIO(file_in_memory), data_only=True)

        try:
            self.run_fast_consistency_checks(wb)
        except ParseError as e:
            raise e

        # If we've made it this far, the checks have passed
        return scan_report


class ScanReportCreateSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    editors = serializers.PrimaryKeyRelatedField(
        many=True, queryset=User.objects.all(), required=False
    )
    parent_dataset = serializers.PrimaryKeyRelatedField(
        queryset=Dataset.objects.order_by("name"),
        required=True,
    )
    visibility = serializers.ChoiceField(
        choices=VisibilityChoices.choices, required=True
    )

    class Meta:
        model = ScanReport
        fields = (
            "editors",
            "dataset",
            "parent_dataset",
            "visibility",
        )

    def validate(self, data):
        if request := self.context.get("request"):
            if ds := data.get("parent_dataset"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(ds, request)
                    or has_editorship(ds, request)
                ):
                    raise PermissionDenied(
                        "You must be either an admin or an editor of the parent dataset to add a new scan report to it.",
                    )
            else:
                raise NotFound("Could not find parent dataset.")
        else:
            raise NotFound("Missing request context. Unable to validate scan report.")
        return super().validate(data)


class ScanReportEditSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    def validate_author(self, author):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    """You must be the author of the scan report or an admin of the parent dataset
                    to change this field."""
                )
        return author

    def validate_viewers(self, viewers):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    """You must be the author of the scan report or an admin of the parent dataset
                    to change this field."""
                )
        return viewers

    def validate_editors(self, editors):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    """You must be the author of the scan report or an admin of the parent dataset
                    to change this field."""
                )
        return editors

    class Meta:
        model = ScanReport
        fields = "__all__"


class DatasetViewSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = Dataset
        fields = "__all__"


class DatasetViewSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = Dataset
        fields = (
            "id",
            "name",
            "data_partner",
            "admins",
            "visibility",
            "created_at",
            "hidden",
            "updated_at",
            "projects",
            "viewers",
            "editors",
        )


class DatasetAndDataPartnerViewSerializer(
    DynamicFieldsMixin, serializers.ModelSerializer
):
    data_partner = DataPartnerSerializer(read_only=True)

    class Meta:
        model = Dataset
        fields = (
            "id",
            "name",
            "data_partner",
            "admins",
            "visibility",
            "created_at",
            "hidden",
        )


class DatasetEditSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    def validate_viewers(self, viewers):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    "You must be an admin to change this field."
                )
        return viewers

    def validate_editors(self, editors):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    "You must be an admin to change this field."
                )
        return editors

    def validate_admins(self, admins):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    "You must be an admin to change this field."
                )
        return admins

    def save(self, **kwargs):
        projects = self.context["projects"]

        if self.instance is not None:
            self.instance = self.update(self.instance, self.validated_data)
            return self.instance
        dataset = Dataset.objects.create(**self.validated_data, projects=projects)
        return dataset

    class Meta:
        model = Dataset
        fields = (
            "id",
            "name",
            "data_partner",
            "admins",
            "visibility",
            "created_at",
            "hidden",
            "updated_at",
            "projects",
            "viewers",
            "editors",
        )


class ScanReportTableListSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    def validate(self, data):
        if request := self.context.get("request"):
            if sr := data.get("scan_report"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(sr, request)
                    or has_editorship(sr, request)
                ):
                    raise PermissionDenied(
                        "You must have editor or admin privileges on the scan report to edit its tables.",
                    )
            else:
                raise NotFound("Could not find the scan report for this table.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report table."
            )
        return super().validate(data)

    class Meta:
        model = ScanReportTable
        fields = "__all__"


class ScanReportTableListSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):

    date_event = serializers.SerializerMethodField()
    person_id = serializers.SerializerMethodField()

    def validate(self, data):
        if request := self.context.get("request"):
            if sr := data.get("scan_report"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(sr, request)
                    or has_editorship(sr, request)
                ):
                    raise PermissionDenied(
                        "You must have editor or admin privileges on the scan report to edit its tables.",
                    )
            else:
                raise NotFound("Could not find the scan report for this table.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report table."
            )
        return super().validate(data)

    def get_date_event(self, obj):
        return obj.date_event.name if obj.date_event else None

    def get_person_id(self, obj):
        return obj.person_id.name if obj.person_id else None

    class Meta:
        model = ScanReportTable
        fields = "__all__"


class ScanReportTableEditSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = ScanReportTable
        fields = "__all__"


class ScanReportFieldListSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    name = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )
    description_column = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )

    def validate(self, data):
        if request := self.context.get("request"):
            if srt := data.get("scan_report_table"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(srt, request)
                    or has_editorship(srt, request)
                ):
                    raise PermissionDenied(
                        "You must have editor or admin privileges on the scan report to edit its fields.",
                    )
            else:
                raise NotFound("Could not find the scan report table for this field.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report field."
            )
        return super().validate(data)

    class Meta:
        model = ScanReportField
        fields = "__all__"


class ScanReportFieldListSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):
    name = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )
    description_column = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )

    def validate(self, data):
        if request := self.context.get("request"):
            if srt := data.get("scan_report_table"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(srt, request)
                    or has_editorship(srt, request)
                ):
                    raise PermissionDenied(
                        "You must have editor or admin privileges on the scan report to edit its fields.",
                    )
            else:
                raise NotFound("Could not find the scan report table for this field.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report field."
            )
        return super().validate(data)

    class Meta:
        model = ScanReportField
        fields = "__all__"


class ScanReportFieldEditSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    name = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )
    description_column = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )

    class Meta:
        model = ScanReportField
        fields = "__all__"


class ScanReportValueViewSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    value = serializers.CharField(
        max_length=128, allow_blank=True, trim_whitespace=False
    )

    def validate(self, data):
        if request := self.context.get("request"):
            if srf := data.get("scan_report_field"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(srf, request)
                    or has_editorship(srf, request)
                ):
                    raise PermissionDenied(
                        "You must have editor or admin privileges on the scan report to edit its values.",
                    )
            else:
                raise NotFound("Could not find the scan report field for this value.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report value."
            )
        return super().validate(data)

    class Meta:
        model = ScanReportValue
        fields = "__all__"


class ScanReportValueViewSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):
    value = serializers.CharField(
        max_length=128, allow_blank=True, trim_whitespace=False
    )

    def validate(self, data):
        if request := self.context.get("request"):
            if srf := data.get("scan_report_field"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(srf, request)
                    or has_editorship(srf, request)
                ):
                    raise PermissionDenied(
                        "You must have editor or admin privileges on the scan report to edit its values.",
                    )
            else:
                raise NotFound("Could not find the scan report field for this value.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report value."
            )
        return super().validate(data)

    class Meta:
        model = ScanReportValue
        fields = "__all__"


class ScanReportValueEditSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    value = serializers.CharField(
        max_length=128, allow_blank=True, trim_whitespace=False
    )

    class Meta:
        model = ScanReportValue
        fields = "__all__"


class ScanReportConceptSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = ScanReportConcept
        fields = "__all__"


class ClassificationSystemSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = ClassificationSystem
        fields = "__all__"


class DataDictionarySerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = DataDictionary
        fields = "__all__"


class OmopFieldSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = OmopField
        fields = "__all__"


class OmopTableSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = OmopTable
        fields = "__all__"


class MappingRuleSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = MappingRule
        fields = "__all__"


class ProjectSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    """
    Serialiser for showing all details of a Project. Use in RetrieveViews
    where User is permitted to view a particular Project.
    """

    class Meta:
        model = Project
        fields = "__all__"


class ProjectNameSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    """
    Serialiser for only showing the names of Projects. Use in non-admin ListViews.
    """

    class Meta:
        model = Project
        fields = ["id", "name", "members"]


class ProjectDatasetSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    """
    Serialiser for only showing the names of Projects. Use in non-admin ListViews.
    """

    class Meta:
        model = Project
        fields = ["name", "datasets", "members"]


class GetRulesJSON(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = ScanReport
        fields = "__all__"

    def to_representation(self, scan_report):
        qs = MappingRule.objects.filter(scan_report=scan_report)
        rules = get_mapping_rules_json(qs)
        return rules


class GetRulesAnalysis(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = ScanReport
        fields = "__all__"

    def to_representation(self, scan_report):
        analysis = analyse_concepts(scan_report.id)
        return analysis


class ContentTypeSerializer(serializers.Serializer):
    """
    Serializes the content type name.

    Args:
        self: The instance of the class.

    Attributes:
        type_name: The serialized content type name.

    """

    type_name = serializers.CharField(max_length=100)
