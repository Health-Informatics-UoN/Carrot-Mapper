from django import forms
from django.contrib.auth import password_validation
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.forms.models import ModelChoiceField

from mapping.models import (DataPartner,
                            DocumentFile, DocumentType,
                            ScanReportField, ScanReport)
import openpyxl
from io import BytesIO


class ShowNameChoiceField(ModelChoiceField):
    def label_from_instance(self, obj):
        if obj.__class__.__name__=='Document':
            return str(obj.data_partner.name)
        return obj.name


class ScanReportForm(forms.Form):
    data_partner = ShowNameChoiceField(
        label="Data Partner",
        queryset=DataPartner.objects.order_by('name'),
        widget=forms.Select(attrs={"class": "form-control"}),
    )
    dataset = forms.CharField(
        label="Dataset name", widget=forms.TextInput(attrs={"class": "form-control"})
    )
    scan_report_file = forms.FileField(
        label="WhiteRabbit ScanReport",
        widget=forms.FileInput(attrs={"class": "form-control"}),
    )

    data_dictionary_file = forms.FileField(
        label="Data Dictionary",
        widget=forms.FileInput(attrs={"class": "form-control"}),
        required=False
    )

    class Meta:
        model = ScanReport
        fields = ('data_partner', 'dataset' , 'scan_report_file')

 
    def clean_data_dictionary_file(self):

        data_dictionary = self.cleaned_data.get("data_dictionary_file")

        if data_dictionary is None:
            return data_dictionary

        if not str(data_dictionary).endswith('.csv'):
            raise ValidationError( "You have attempted to upload a data dictionary which is not in CSV format. Please upload a .csv file.")
        
        return data_dictionary

    def run_fast_consistency_checks(self, wb):
        """
        Runs a number of consistency checks on the provided workbook. The aim is to
        return quickly if there is an issue with the data, and provide feedback to the
        user so they can fix the issue.
        """

        # Get the first sheet 'Field Overview'
        fo_ws = wb.worksheets[0]

        # Grab the scan report columns from the first worksheet
        # Define what the column headings should be
        source_headers = [header.value for header in fo_ws[1]]

        expected_headers = ['Table', 'Field', 'Description', 'Type', 'Max length',
                            'N rows', 'N rows checked', 'Fraction empty',
                            'N unique values', 'Fraction unique']

        # Check if source headers match the expected headers. Allow unexpected
        # headers after these. This means old Scan Reports with Flag and Classification
        # columns will be handled cleanly.
        if not source_headers[:10] == expected_headers:
            raise ValidationError(f"Please check the following columns exist in the "
                                  f"Scan Report (Field Overview sheet) in this order: "
                                  f"Table, Field, Description, Type, Max length, "
                                  f"N rows, N rows checked, Fraction empty, "
                                  f"N unique values, Fraction unique. "
                                  f"You provided \n{source_headers[:10]}")

        # Check tables are correctly separated in FO - a single empty line between each
        # table
        cell_above = fo_ws['A'][1]
        for cell in fo_ws['A'][1:]:
            if (cell.value != cell_above.value and
                cell.value != '' and
                cell_above.value != '') or \
                    (cell.value == '' and cell_above.value == ''):
                raise ValidationError(f"At {cell}, tables in Field Overview table are "
                                      f"not correctly separated by a single line. "
                                      f"Note: There should be no separator line "
                                      f"between the header row and the first row of "
                                      f"the first table.")
            cell_above = cell

        # Now that we're happy that the FO sheet is correctly formatted, we can move
        # on to comparing its contents to the sheets

        # Check tables in FO match supplied sheets
        table_names = list(
            set(cell.value for cell in fo_ws['A'][1:] if cell.value != '')
        )
        expected_sheetnames = table_names + ['Field Overview', 'Table Overview', '_']
        if sorted(wb.sheetnames) != sorted(expected_sheetnames):
            sheets_only = set(wb.sheetnames).difference(expected_sheetnames)
            fo_only = set(expected_sheetnames).difference(wb.sheetnames)
            error_text = f"Tables in Field Overview sheet do not match the sheets supplied."
            if sheets_only:
                error_text += f"{sheets_only} are sheets that do not have matching " \
                              f"entries in first column of the Field Overview sheet. "
            if fo_only:
                error_text += f"{fo_only} are table names in first column of Field " \
                              f"Overview sheet but do not have matching sheets supplied."
            raise ValidationError(error_text)

        # Loop over the rows, and for each table, once we reach the end of the table,
        # compare the fields provided with the fields in the associated sheet
        current_table_fields = []
        for row in fo_ws.iter_rows(min_row=2):
            # Loop over rows, collecting all fields in each table in turn
            if row[0].value == '':
                # We're at the end of the table, so process
                # Get all field names from the associated sheet, by grabbing the first
                # row, and then grabbing every second column value (because the
                # alternate columns should be 'Frequency'
                table_sheet_fields = set([cell.value
                                          for cell in next(wb[current_table_name].rows)
                                          ][::2]
                                         )
                if sorted(table_sheet_fields) != sorted(current_table_fields):
                    sheet_only = set(table_sheet_fields).difference(
                        current_table_fields)
                    fo_only = set(current_table_fields).difference(table_sheet_fields)
                    error_text = f"Fields in Field Overview against table " \
                                 f"{current_table_name} do not match fields in the " \
                                 f"associated sheet. "
                    if sheet_only:
                        error_text += f"{sheet_only} exist in the " \
                                      f"'{current_table_name}' " \
                                      f"sheet but there are no matching entries in " \
                                      f"the second column of the Field Overview " \
                                      f"sheet in the rows associated to the table " \
                                      f"'{current_table_name}'. "
                    if fo_only:
                        error_text += f"{fo_only} exist in second column of Field " \
                                      f"Overview sheet against the table " \
                                      f"'{current_table_name}' but there are no " \
                                      f"matching column names in the associated " \
                                      f"sheet '{current_table_name}'."
                    raise ValidationError(error_text)

                # Reset the list of fields associated to this table as we iterate down
                # the FO sheet.
                current_table_fields = []
            else:
                # Update current list of field names and the current table name - we can
                # trust the table name not to change in this case due to the earlier
                # check for empty lines between tables in the FO sheet.
                current_table_fields.append(row[1].value)
                current_table_name = row[0].value

        return True

    def clean_scan_report_file(self):

        scan_report = self.cleaned_data.get("scan_report_file")

        if not str(scan_report).endswith('.xlsx'):
            raise ValidationError("You have attempted to upload a scan report which "
                                  "is not in XLSX format. Please upload a .xlsx file.")

        # Load in the Excel sheet, grab the first workbook
        file_in_memory = scan_report.read()
        wb = openpyxl.load_workbook(filename=BytesIO(file_in_memory), data_only=True)

        self.run_fast_consistency_checks(wb)

        return scan_report


class UserCreateForm(UserCreationForm):
    email = forms.EmailField(
        required=True, label="Email", error_messages={"exists": "Oops"}
    )

    class Meta:
        model = User
        fields = ("username", "email", "password1", "password2")

    def save(self, commit=True):
        user = super(UserCreateForm, self).save(commit=False)
        user.email = self.cleaned_data["email"]
        if commit:
            user.save()
        return user

    def clean_email(self):
        if User.objects.filter(email=self.cleaned_data["email"]).exists():
            raise ValidationError(self.fields["email"].error_messages["exists"])
        return self.cleaned_data["email"]


class PasswordChangeForm(forms.Form):
    old_password = forms.CharField(
        label="Old Password",
        widget=forms.PasswordInput(
            attrs={"autocomplete": "current-password", "autofocus": True}
        ),
    )
    new_password1 = forms.CharField(
        label=("New password"),
        widget=forms.PasswordInput(attrs={"autocomplete": "new-password"}),
        validators=[password_validation.validate_password]

    )

    new_password2 = forms.CharField(
        label=("Confirm New password"),
        widget=forms.PasswordInput(attrs={"autocomplete": "new-password"}),
        validators=[password_validation.validate_password]
    )


class DocumentForm(forms.Form):
    
    data_partner = ShowNameChoiceField(
        label="Data Partner",
        queryset=DataPartner.objects.order_by("name"),
        widget=forms.Select(attrs={"class": "form-control"}),
    )

    document_type = ShowNameChoiceField(
        label="Type",
        queryset=DocumentType.objects.order_by("name"),
        widget=forms.Select(attrs={"class": "form-control"}),
    )

    description = forms.CharField(
        label="Description", widget=forms.TextInput(attrs={"class": "form-control"})
    )

    document_file = forms.FileField(
        label="File", widget=forms.FileInput(attrs={"class": "form-control"})
    )

    def clean_document_file(self):
        if str((self.cleaned_data['document_type']).name).lower()=='data dictionary':
            try:
                data_dictionary_csv = self.cleaned_data['document_file'].read().decode("utf-8-sig").splitlines()[0]
                header = data_dictionary_csv.split(',')
                column_names = ["TableName", "FieldName", "FieldDescription", "Value", "ValueDescription"]

                if set(column_names) == set(header):
                    return self.cleaned_data['document_file']
                else:
                    raise (forms.ValidationError("Please check your column names in your data dictionary"))
            except: 
                raise (forms.ValidationError("Data Dictionary must be .csv file"))
        else:
            return self.cleaned_data['document_file']


class DocumentFileForm(forms.Form):
    document_file = forms.FileField(
        label="Document File", widget=forms.FileInput(attrs={"class": "form-control"})
    )
    document_type = ShowNameChoiceField(
        label="Type",
        queryset=DocumentType.objects.order_by("name"),
        widget=forms.Select(attrs={"class": "form-control"}),
    )
    description = forms.CharField(
        label="Document Description",
        widget=forms.TextInput(attrs={"class": "form-control"}),
    )

    def clean_document_type(self):
        print(self.cleaned_data['document_type'].name)
        if str((self.cleaned_data['document_type']).name).lower()=='data dictionary':
            try:
                data_dictionary_csv = self.cleaned_data['document_file'].read().decode("utf-8-sig").splitlines()[0]
                header = data_dictionary_csv.split(',')
                column_names = ["TableName", "FieldName", "FieldDescription", "Value", "ValueDescription"]

                if set(column_names) == set(header):
                    return self.cleaned_data['document_file']
                else:
                    raise (forms.ValidationError("Please check your column names in your data dictionary"))
            except: 
                raise (forms.ValidationError("Data Dictionary must be .csv file"))
        else:
            return self.cleaned_data['document_file']
      
        
class DictionarySelectForm(forms.Form):
    document = forms.ModelChoiceField(label="Data Dictionary Document",
                                      queryset=DocumentFile.objects.filter(status__icontains="LIVE"),
                                      to_field_name="document")


class ScanReportAssertionForm(forms.Form):
    negative_assertion=forms.CharField(
         label="Negative Assertions",
         widget=forms.TextInput(attrs={"class": "form-control"}),
     )    


class NLPForm(forms.Form):
    user_string = forms.CharField(
        label = ""
    )


class ScanReportFieldConceptForm(forms.Form):
    scan_report_field_id=forms.IntegerField(

    )

    concept_id=forms.IntegerField(

    )


class ScanReportValueConceptForm(forms.Form):
    scan_report_value_id=forms.IntegerField(

    )

    concept_id=forms.IntegerField(

    )


class ScanReportFieldForm(forms.ModelForm):
    class Meta:
        model = ScanReportField
        fields = (
            #"is_patient_id",
            #"is_date_event",
            "is_ignore",
            "pass_from_source",
            "description_column"
            )
        widgets = {
            'description_column': forms.TextInput(attrs={
            'class': u'form-control'})
        }
