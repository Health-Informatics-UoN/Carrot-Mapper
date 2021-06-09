import logging
import json
import ast
import azure.functions as func
from azure.storage.blob import BlockBlobService
from io import BytesIO
import requests
import openpyxl
from datetime import datetime
import os

def process_scan_report_sheet_table(sheet):
    """
    This function extracts the
    data into the format below.

    -- Example Table Sheet CSV --
    column a,frequency,column c,frequency
    apple,      20,     orange,      5
    banana,     3,      plantain,    50
    
    --

    -- output --
    column a, apple, 20
    column c, orange, 5
    column a, banana, 3
    column c, plantain, 50
    --
    """
    results = []
    # Get max number of columns in the sheet
    max_column=sheet.max_column
    # Set a column index (openpyxl index at 1 instead of 0)
    col_idx=1
    
    for row_idx,row_cell in enumerate(sheet.iter_rows(min_row=1,max_col=max_column),start = 1):
  
    # Works through pairs of value/frequency columns
        for column_idx,cell in enumerate(row_cell):
          if cell.value:
            if (column_idx) % 2 == 0:
            # As we move down rows, checks that there's data there
            # This is required b/c value/frequency col pairs differ
            # in the number of rows
              if sheet.cell(row=row_idx,column=column_idx+1).value == "" and sheet.cell(row=row_idx,column=column_idx+2).value == "":
                continue
            # Append to Results as (Field Name,Value,Frequency)
              results.append(
                  (sheet.cell(row=1,column=column_idx+1).value,sheet.cell(row=row_idx,column=column_idx+1).value, sheet.cell(row=row_idx,column=column_idx+2).value)
              )
    col_idx=col_idx+1

    return results


def main(msg: func.QueueMessage):
    logging.info('Python queue trigger function processed a queue item.')
    # Get message from queue
    message = json.dumps({
        'id': msg.id,
        'body': msg.get_body().decode('utf-8'),
        'expiration_time': (msg.expiration_time.isoformat()
                            if msg.expiration_time else None),
        'insertion_time': (msg.insertion_time.isoformat()
                           if msg.insertion_time else None),
        'time_next_visible': (msg.time_next_visible.isoformat()
                              if msg.time_next_visible else None),
        'pop_receipt': msg.pop_receipt,
        'dequeue_count': msg.dequeue_count
    })
    message=json.loads(message)
    body=ast.literal_eval(message["body"])
    filename=body["blob_name"]
    # Write File saved in Blob storage to BytesIO stream
    with BytesIO() as input_blob:
        # Connect to Blob Service
        blob = BlockBlobService(connection_string=os.environ.get('coconnectstoragedev_STORAGE'))
        # Download Scan Report .xlsx file as a ByteIO() stream
        blob.get_blob_to_stream(container_name='photos', blob_name=filename,stream=input_blob)
        input_blob.seek(0)

        # Set up API parameters:
        api_url="http://localhost:8080/api/"
        headers={"Content-type": "application/json", "charset":"utf-8"}

        # Load ByteIO() file in a openpyxl workbook
        wb = openpyxl.load_workbook(input_blob,data_only=True)
        # Get the first sheet 'Field Overview',
        # to populate ScanReportTable & ScanReportField models
        ws=wb.worksheets[0]
        
        table_names=[]
        # Skip header with min_row=2
        for i,row_cell in enumerate(ws.iter_rows(min_row=2),start = 2):
            # If the value in the first column (i.e. the Table Col) is not blank;
            # Save the table name as a new entry in the model ScanReportTable
            # Checks for blank b/c White Rabbit seperates tables with blank row
            for cell in row_cell:
                if cell.value:
                    name=ws.cell(row=i,column=1).value
                    if name not in table_names:
                        table_names.append(name)

        """
        For each table create a scan_report_table entry,
        Append entry to data[] list,
        Create JSON array with all the entries, 
        Send POST request to API with JSON as input,
        Save the response data(table IDs)
        """
        table_ids = []
        field_ids=[]
        field_names=[]
        data=[]
        for table in range(len(table_names)):

            print('WORKING ON TABLE >>> ', table)
            # Truncate table names because sheet names are truncated to 31 characters in Excel
            table_names[table]=table_names[table][:31]
        
            # Create ScanReportTable entry
            # Link to scan report using ID from the queue message
            scan_report_table_entry={
                    "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                    "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                    "name": table_names[table],
                    "scan_report":str(body['scan_report_id']),
                    "person_id": None,
                    "birth_date": None,
                    "measurement_date": None,
                    "condition_date": None,
                    "observation_date": None
            }
            # Append to list
            data.append(scan_report_table_entry)
        # Create JSON array
        json_data=json.dumps(data)
        # POST request to scanreporttables
        response = requests.post("{}scanreporttables/".format(api_url), data=json_data,headers=headers)
        print('TABLE SAVE STATUS >>>', response.status_code)
        # Load the result of the post request,
        # Save the table ids that were generated from the POST method
        response=json.loads(response.content.decode("utf-8"))
        for element in range(len(response)):
            table_ids.append(response[element]['id'])
        print('TABLE IDs', table_ids)

        """
        For each Column in Field Overview create an entry for scan_report_field,
        If there is an empty row then write to the next table
        Append entry to data[] list,
        Create JSON array with all the field entries, 
        Send POST request to API with JSON as input,
        Save the response data(table IDs)
        """
        idx=0
        data=[]
        for i,row_cell in enumerate(ws.iter_rows(min_row=2),start = 2):
            
            # Create ScanReportField entry
            scan_report_field_entry={
                "scan_report_table":table_ids[idx],
                "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "name": ws.cell(row=i,column=2).value,
                "description_column":str("empty"),
                "type_column":str(ws.cell(row=i,column=4).value),
                "max_length":ws.cell(row=i,column=5).value,
                "nrows":ws.cell(row=i,column=6).value,
                "nrows_checked":ws.cell(row=i,column=7).value,
                "fraction_empty":round(ws.cell(row=i,column=8).value,2),
                "nunique_values":ws.cell(row=i,column=9).value,
                "fraction_unique":round(ws.cell(row=i,column=10).value,2),
                "flag_column":str(ws.cell(row=i,column=11).value),
                "ignore_column":None,
                "is_birth_date":False,
                "is_patient_id":False,
                "is_date_event":False,
                "is_ignore":False,
                "pass_from_source":True,
                "classification_system":str(ws.cell(row=i,column=12).value),
                "date_type": "",
                "concept_id": -1,
                "field_description": None,
            }
            # Append each entry to list
            data.append(scan_report_field_entry)
            # If there was an empty row remove it from list,
            # Switch to writing to the next table
            if not any(cell.value for cell in row_cell):
                idx=idx+1
                data.pop()
        # Create JSON array
        json_data=json.dumps(data)
        # POST request
        response = requests.post("{}scanreportfields/".format(api_url), data=json_data,headers=headers)
        print('FIELD SAVE STATUS >>>', response.status_code)
        # Load result from the response,
        # Save generated field ids, and the corresponding name
        response=json.loads(response.content.decode("utf-8"))
        
        for element in range(len(response)):
            field_ids.append(str(response[element].get('id',None)))
            field_names.append(str(response[element].get('name',None)))
        
        # Fix for empty row (empty row was being returned as [] element)
        # Need to remove for dict(zip()) to work
        for name in field_names:
            if name.startswith("[") or None:
              field_names.remove(name)
        for id in field_ids:
            if id=="None": field_ids.remove(id)
        # Create a dictionary with field names and field ids 
        # as key value pairs
        # e.g ("Field Name":<Field ID>)
        namexids=dict(zip(field_names,field_ids))

        # For sheets past the first two in the Scan Report
        # i.e. all 'data' sheets that are not Field Overview and Table Overview
        data=[]
        for idxsheet, sheet in enumerate(wb.worksheets):
            if idxsheet<2:
                continue
        
            # skip these sheets at the end of the scan report
            if sheet.title=="_" or (sheet.title.startswith("HTA")):
                continue
            print("WORKING ON SHEET>>>",sheet.title)
            # Get value,frequency for each field
            results=process_scan_report_sheet_table(sheet)
            """
            For every result of process_scan_report_sheet_table,
            Save the current name,value,frequency
            Create ScanReportValue entry,
            Append to data[] list,
            Create JSON array with all the field entries, 
            Send POST request to API with JSON as input
            """
            for result in range(len(results)):
                
                name=str(results[result][0])
                value=str(results[result][1])
                frequency=results[result][2]
                # Set frequency to 0 when empty since the model expects an int
                if (frequency is None) or (frequency==""):
                    frequency=0    
                # If we are not on the first row:
                if name!=value:
                    # Create a ScanReportValue entry
                    scan_report_value_entry={
                        "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                        "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                        "value": value,
                        "frequency": int(frequency),
                        "conceptID": -1,
                        "value_description": None,
                        "scan_report_field":namexids[name]
                    }
                    # Append to list
                    data.append(scan_report_value_entry)
    # Create JSON array
    json_data=json.dumps(data)
    # POST request
    response = requests.post("{}scanreportvalues/".format(api_url), data=json_data,headers=headers)
    print('VALUE SAVE STATUS >>>', response.status_code)

    logging.info(body['blob_name'])
