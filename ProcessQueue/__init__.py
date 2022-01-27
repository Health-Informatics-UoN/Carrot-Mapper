import logging
import json
import ast
import azure.functions as func
from azure.storage.blob import BlobServiceClient
from io import BytesIO
import requests
import openpyxl
from datetime import datetime
import os
import csv
import psutil
import httpx
import asyncio

from requests.models import HTTPError
from collections import defaultdict

from shared_code import omop_helpers

# import memory_profiler
# root_logger = logging.getLogger()
# root_logger.handlers[0].setFormatter(logging.Formatter("%(name)s: %(message)s"))
# profiler_logstream = memory_profiler.LogFile('memory_profiler_logs', True)


# Agreed vocabs that are accepted for lookup/conversion
# The Data Team decide what vocabs are accepted.
# Add more as necessary by appending the list
vocabs = [
    "ABMS",
    "ATC",
    "HCPCS",
    "HES Specialty",
    "ICD10",
    "ICD10CM",
    "ICD10PCS",
    "ICD9CM",
    "ICD9Proc",
    "LOINC",
    "NDC",
    "NUCC",
    "OMOP Extension",
    "OSM",
    "PHDSC",
    "Read",
    "RxNorm",
    "RxNorm Extension",
    "SNOMED",
    "SPL",
    "UCUM",
    "UK Biobank",
]


# @memory_profiler.profile(stream=profiler_logstream)
def process_scan_report_sheet_table(sheet):
    """
    This function extracts the
    data into the format below.

    -- Example Table Sheet CSV --
    a,   frequency,          b, frequency
    apple,      20,     orange,         5
    banana,      3,   plantain,        50
    pear,       12,         '',        ''

    --

    -- output --
    [(a,    apple, 20),
     (a,   banana,  3),
     (a,     pear, 12),
     (b,   orange,  5),
     (b, plantain, 50)]
    --
    """
    print(
        "Start process_scan_report_sheet_table",
        datetime.utcnow().strftime("%H:%M:%S.%fZ"),
        flush=True,
    )
    # Get header entries (skipping every second column which is just 'Frequency')
    # So headers = ['a', 'b']
    first_row = sheet[1]
    headers = [cell.value for cell in first_row[::2]]

    # Set up an empty defaultdict, and fill it with one entry per header (i.e. one
    # per column)
    # Append each entry's value with the tuple (value, frequency) so that we end up
    # with each entry containing one tuple per non-empty entry in the column.
    #
    # This will give us
    #
    # ordereddict({'a': [('apple', 20), ('banana', 3), ('pear', 12)],
    #              'b': [('orange', 5), ('plantain', 50)]})

    d = defaultdict(list)
    # Iterate over all rows beyond the header - use the number of headers*2 to
    # set the maximum column rather than relying on sheet.max_col as this is not
    # always reliably updated by Excel etc.
    for row in sheet.iter_rows(
        min_col=1,
        max_col=len(headers) * 2,
        min_row=2,
        max_row=sheet.max_row,
        values_only=True,
    ):
        # Set boolean to track whether we hit a blank row for early exit below.
        this_row_empty = True
        # Iterate across the pairs of cells in the row. If the pair is non-empty,
        # then add it to the relevant dict entry.
        for (header, cell, freq) in zip(headers, row[::2], row[1::2]):
            if cell != "" or freq != "":
                d[header].append((str(cell), freq))
                this_row_empty = False
        # This will trigger if we hit a row that is entirely empty. Short-circuit
        # to exit early here - this saves us from situations where sheet.max_row is
        # incorrectly set (too large)
        if this_row_empty:
            break

    print(
        "Finish process_scan_report_sheet_table",
        datetime.utcnow().strftime("%H:%M:%S.%fZ"),
        flush=True,
    )
    return d


def default_zero(input):
    """
    Helper function that returns the input, replacing anything Falsey
    (such as Nones or empty strings) with 0.0.
    """
    return round(input if input else 0.0, 2)


def perform_chunking(entries_to_post):
    """
    This expects a list of dicts, and returns a list of lists of lists of dicts,
    where the maximum length of each list of dicts, under JSONification,
    is less than max_chars, and the length of each list of lists of dicts is chunk_size
    """
    max_chars = (
        int(os.environ.get("PAGE_MAX_CHARS"))
        if os.environ.get("PAGE_MAX_CHARS")
        else 10000
    )
    chunk_size = (
        int(os.environ.get("CHUNK_SIZE")) if os.environ.get("CHUNK_SIZE") else 6
    )

    chunked_entries_to_post = []
    this_page = []
    this_chunk = []
    page_no = 0
    for entry in entries_to_post:
        # If the current page won't be overfull, add the entry to the current page
        if len(json.dumps(this_page)) + len(json.dumps(entry)) < max_chars:
            this_page.append(entry)
        # Otherwise, this page should be added to the current chunk.
        else:
            this_chunk.append(this_page)
            page_no += 1
            # Now check for a full chunk. If full, then add this chunk to the list of chunks.
            if page_no % chunk_size == 0:
                # append the chunk to the list of chunks, then reset the chunk to empty
                chunked_entries_to_post.append(this_chunk)
                this_chunk = []
            # Now add the entry that would have over-filled the page.
            this_page = [entry]
    # After all entries are added, check for a half-filled page, and if present add it to the list of pages
    if this_page:
        this_chunk.append(this_page)
    # Similarly, if a chunk ends up half-filled, add it to thelist of chunks
    if this_chunk:
        chunked_entries_to_post.append(this_chunk)

    return chunked_entries_to_post


# @memory_profiler.profile(stream=profiler_logstream)
def paginate(entries_to_post, max_chars=None, other=""):
    """
    This expects a list of dicts, and returns a list of lists of dicts,
    where the maximum length of each list of dicts, under JSONification,
    is less than max_chars
    """
    if not max_chars:
        max_chars = (
            int(os.environ.get("PAGE_MAX_CHARS"))
            if os.environ.get("PAGE_MAX_CHARS")
            else 10000
        )
    max_chars = max_chars - len(other)

    paginated_entries_to_post = []
    this_page = []
    for entry in entries_to_post:
        # If the current page won't be overfull, add the entry to the current page
        if len(json.dumps(this_page)) + len(json.dumps(entry)) < max_chars:
            this_page.append(entry)
        else:
            # Otherwise, this page should be added to the list of pages.
            paginated_entries_to_post.append(this_page)
            # Now add the entry that would have over-filled the page.
            this_page = [entry]

    # After all entries are added, check for a half-filled page, and if present add it to the list of pages
    if this_page:
        paginated_entries_to_post.append(this_page)
    return paginated_entries_to_post


# @memory_profiler.profile(stream=profiler_logstream)
def startup(msg):
    logging.info("Python queue trigger function processed a queue item.")
    print("RAM memory % used:", psutil.virtual_memory())
    print(datetime.utcnow().strftime("%H:%M:%S.%fZ"))
    # Set up ccom API parameters:
    api_url = os.environ.get("APP_URL") + "api/"
    headers = {
        "Content-type": "application/json",
        "charset": "utf-8",
        "Authorization": "Token {}".format(os.environ.get("AZ_FUNCTION_KEY")),
    }

    # Get message from queue
    message = {
        "id": msg.id,
        "body": msg.get_body().decode("utf-8"),
        "expiration_time": (
            msg.expiration_time.isoformat() if msg.expiration_time else None
        ),
        "insertion_time": (
            msg.insertion_time.isoformat() if msg.insertion_time else None
        ),
        "time_next_visible": (
            msg.time_next_visible.isoformat() if msg.time_next_visible else None
        ),
        "pop_receipt": msg.pop_receipt,
        "dequeue_count": msg.dequeue_count,
    }

    print("message:", type(message), message)
    # Grab message body from storage queues,
    # extract filenames for scan reports and dictionaries
    # print("body 1:", type(message["body"]), message["body"])
    body = json.loads(message["body"])
    # print("body 2:", type(body), body)
    scan_report_blob = body["scan_report_blob"]
    data_dictionary_blob = body["data_dictionary_blob"]

    print("MESSAGE BODY >>>", body)

    # If the message has been dequeued for a second time, then the upload has failed.
    # Patch the name of the dataset to make it clear that it has failed,
    # set the status to 'Upload Failed', and then stop.
    print("dequeue_count", msg.dequeue_count)
    scan_report_id = body["scan_report_id"]
    if msg.dequeue_count == 2:
        process_failure(api_url, scan_report_id, headers)

    if msg.dequeue_count > 1:
        raise Exception("dequeue_count > 1")

    # Otherwise, this must be the first time we've seen this message. Proceed.
    return api_url, headers, scan_report_blob, data_dictionary_blob, scan_report_id


def process_failure(api_url, scan_report_id, headers):
    scan_report_fetched_data = requests.get(
        url=f"{api_url}scanreports/{scan_report_id}/",
        headers=headers,
    )

    scan_report_fetched_data = json.loads(
        scan_report_fetched_data.content.decode("utf-8")
    )

    json_data = json.dumps(
        {
            "dataset": f"FAILED: {scan_report_fetched_data['dataset']}",
            "status": "UPFAILE",
        }
    )

    failure_response = requests.patch(
        url=f"{api_url}scanreports/{scan_report_id}/", data=json_data, headers=headers
    )


def paginate_chars(entries_to_post, other):
    """
    This expects a list of dicts, and returns a list of lists of dicts,
    where the maximum length of each list of dicts, under JSONification,
    is less than max_chars
    """
    max_chars = 2000 - len(other)

    paginated_entries_to_post = []
    this_page = []
    for entry in entries_to_post:
        # If the current page won't be overfull, add the entry to the current page
        if len(json.dumps(this_page)) + len(json.dumps(entry)) < max_chars:
            this_page.append(entry)
        else:
            # Otherwise, this page should be added to the list of pages.
            paginated_entries_to_post.append(this_page)
            # Now add the entry that would have over-filled the page.
            this_page = [entry]

    # After all entries are added, check for a half-filled page, and if present add it to the list of pages
    if this_page:
        paginated_entries_to_post.append(this_page)
    return paginated_entries_to_post


def flatten(arr):
    """
    This expects a list of lists and returns a flattened list
    """
    newArr = [item for sublist in arr for item in sublist]
    return newArr


def reuse_existing_field_concepts(new_fields_map, content_type, api_url, headers):
    """
    This expects a dict of field names to ids which have been generated in a newly uploaded
    scanreport, and content_type 15. It creates new concepts associated to any
    field that matches the name of an existing field with an associated concept.
    """
    # Gets all scan report concepts that are for the type field (or content type which should be field)
    get_field_concept_ids = requests.get(
        url=f"{api_url}scanreportconceptsfilter/?content_type={content_type}",
        headers=headers,
    )
    field_concept_ids = json.loads(get_field_concept_ids.content.decode("utf-8"))
    # Creates a dictionary that maps field id's to scan report concept id's
    field_id_to_concept_map = {
        element.get("object_id", None): str(element.get("concept", None))
        for element in field_concept_ids
    }
    print("FIELD TO CONCEPT MAP DICT", field_id_to_concept_map)
    # creates a list of field ids from fields that already exist
    existing_ids = list(field_id_to_concept_map.keys())
    # create list of field names from list of newly generated fields
    new_fields_names_string = ",".join(map(str, list(new_fields_map.keys())))
    # paginate the field id's variable so that get request does not exceed character limit
    paginated_ids = paginate(existing_ids, 2000, new_fields_names_string)
    # for each list in paginated id's, get scanreport fields that match any of the given id's
    # and matches any of the newly generated names
    fields = []
    for ids in paginated_ids:
        ids_to_get = ",".join(map(str, ids))
        get_field_names = requests.get(
            url=f"{api_url}scanreportfieldsfilter/?id__in={ids_to_get}&name__in={new_fields_names_string}",
            headers=headers,
        )
        fields.append(json.loads(get_field_names.content.decode("utf-8")))
    fields = flatten(fields)
    print("FIELDS", fields)

    # get a list of table ids for all the fields that have matching names
    table_ids = set([item["scan_report_table"] for item in fields])
    # get tables from list of table ids
    paginated_table_ids = paginate(table_ids, 2000, "")
    tables = []
    for ids in paginated_table_ids:
        ids_to_get = ",".join(map(str, ids))

        get_field_tables = requests.get(
            url=f"{api_url}scanreporttablesfilter/?id__in={ids_to_get}",
            headers=headers,
        )
        tables.append(json.loads(get_field_tables.content.decode("utf-8")))
    tables = flatten(tables)
    print("TABLES", tables)
    # map table id's to scanreport id

    # get all scanreports to be used to check active scan reports
    get_scan_reports = requests.get(
        url=f"{api_url}scanreports/",
        headers=headers,
    )
    # get active scanreports and map them to fields. Remove any fields in archived reports
    scanreports = json.loads(get_scan_reports.content.decode("utf-8"))
    active_reports = [
        str(item["id"])
        for item in scanreports
        if item["hidden"] == False and item["status"] == "COMPLET"
    ]
    # active reports is list of report ids that are not archived
    table_id_to_active_scanreport_map = {
        str(element["id"]): str(element["scan_report"])
        for element in tables
        if str(element["scan_report"]) in active_reports
    }
    # map field id to active scan report id. (only store field ids that correspond to an active scan report)
    field_id_to_active_scanreport_map = {
        str(element["id"]): table_id_to_active_scanreport_map[
            str(element["scan_report_table"])
        ]
        for element in fields
        if str(element["scan_report_table"]) in table_id_to_active_scanreport_map
    }
    # filter fields to only include fields that are from active scan reports
    fields = [
        item for item in fields if str(item["id"]) in field_id_to_active_scanreport_map
    ]
    print("FILTERED FIELDS", fields)

    existing_mappings = [
        {
            "name": field["name"],
            "concept": field_id_to_concept_map[field["id"]],
            "id": field["id"],
        }
        for field in fields
    ]
    print("EXISTING MAPPINGS", existing_mappings)
    field_name_to_id_map = {}
    for name in list(new_fields_map.keys()):
        mappings_matching_field_name = [
            mapping for mapping in existing_mappings if mapping["name"] == name
        ]
        target_concept_ids = set(
            [mapping["concept"] for mapping in mappings_matching_field_name]
        )
        target_field_id = set(
            [mapping["id"] for mapping in mappings_matching_field_name]
        )
        if len(target_concept_ids) == 1:
            field_name_to_id_map[str(name)] = str(target_field_id.pop())

    # replace field_name_to_id_map with field name to concept id map
    # field_name_to_concept_id_map = { element.key: field_id_to_concept_map[int(element.value)] for element in field_name_to_id_map }

    print("FIELD NAME TO ID MAP", field_name_to_id_map)
    concepts_to_post = []
    concept_response_content = []
    print("NAME IDS", new_fields_map.keys())

    for name, id in new_fields_map.items():
        try:
            link_id = field_name_to_id_map[name]
            concept_id = field_id_to_concept_map[int(link_id)]

            print(
                f"Found field with id: {link_id} with exsting concept mapping: {concept_id} which matches new field id: {id}"
            )
            # Create ScanReportConcept entry for copying over the concept
            concept_entry = {
                "nlp_entity": None,
                "nlp_entity_type": None,
                "nlp_confidence": None,
                "nlp_vocabulary": None,
                "nlp_processed_string": None,
                "concept": concept_id,
                "object_id": id,
                "content_type": content_type,
                "creation_type": "R",
            }
            concepts_to_post.append(concept_entry)
        except KeyError:
            continue
    if concepts_to_post:

        paginated_concepts_to_post = paginate(concepts_to_post)
        concept_response = []
        for concepts_to_post_item in paginated_concepts_to_post:
            get_concept_response = requests.post(
                url=api_url + "scanreportconcepts/",
                headers=headers,
                data=json.dumps(concepts_to_post_item),
            )
            print(
                "CONCEPTS SAVE STATUS >>>",
                get_concept_response.status_code,
                get_concept_response.reason,
                flush=True,
            )
            concept_response.append(
                json.loads(get_concept_response.content.decode("utf-8"))
            )
        concept_content = flatten(concept_response)

        concept_response_content += concept_content

        print("POST concepts all finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))


def reuse_existing_value_concepts(new_values_map, content_type, api_url, headers):
    """
    This expects a dict of value names to ids which have been generated in a newly uploaded scanreport and
    creates new concepts if any matching names are found with existing fields
    """
    # get all scan report concepts with the concept type of values (or content type but should be values)
    get_value_concept_ids = requests.get(
        url=f"{api_url}scanreportconceptsfilter/?content_type={content_type}&fields=object_id,concept",
        headers=headers,
    )
    # create dictionary that maps existing value id's to scan report concepts
    # from the list of existing scan report concepts
    value_concept_ids = json.loads(get_value_concept_ids.content.decode("utf-8"))
    value_id_to_concept_map = {
        str(element.get("object_id", None)): str(element.get("concept", None))
        for element in value_concept_ids
    }
    # create list of names of newly generated values
    new_values_names_list = [value["value"] for value in new_values_map]

    new_paginated_field_ids = paginate(
        [value["scan_report_field"] for value in new_values_map], 2000
    )
    new_fields = []
    for ids in new_paginated_field_ids:
        ids_to_get = ",".join(map(str, ids))
        get_fields = requests.get(
            url=f"{api_url}scanreportfieldsfilter/?id__in={ids_to_get}",
            headers=headers,
        )
        new_fields.append(json.loads(get_fields.content.decode("utf-8")))
    new_fields = flatten(new_fields)

    new_fields_to_name_map = {str(value["id"]): value["name"] for value in new_fields}

    new_values_matching_list = [
        {
            "name": value["value"],
            "description": value["value_description"],
            "field_name": new_fields_to_name_map[str(value["scan_report_field"])],
        }
        for value in new_values_map
    ]

    # create dictionary that maps value names to value ids
    new_value_ids = {str(value["value"]): value["id"] for value in new_values_map}

    # create list of id's from existing values that have a scanreport concept
    existing_ids = list(value_id_to_concept_map.keys())
    # create names string to pass to url from list of names made from newly generated values
    new_value_names_string = ",".join(map(str, new_values_names_list))
    # paginate list of value ids from existing values that have scanreport concepts and
    # use the list to get existing scanreport values that match the list any of the newly generated names
    paginated_ids = paginate(existing_ids, 2000, new_value_names_string)
    print("VALUE names list", new_value_names_string)
    print("VALUE paginated ids", paginated_ids)
    existing_scanreport_values = []
    for ids in paginated_ids:
        ids_to_get = ",".join(map(str, ids))
        get_value_names = requests.get(
            url=f"{api_url}scanreportvaluesfilter/?id__in={ids_to_get}&value__in={new_value_names_string}&fields=id,value,scan_report_field,value_description",
            headers=headers,
        )
        existing_scanreport_values.append(
            json.loads(get_value_names.content.decode("utf-8"))
        )
    existing_scanreport_values = flatten(existing_scanreport_values)
    print("Unfiltered values", existing_scanreport_values)

    # get field ids from values and use to get scan report fields
    field_ids = set([item["scan_report_field"] for item in existing_scanreport_values])
    paginated_field_ids = paginate(field_ids, 2000, "")
    fields = []
    for ids in paginated_field_ids:
        ids_to_get = ",".join(map(str, ids))

        get_value_fields = requests.get(
            url=f"{api_url}scanreportfieldsfilter/?id__in={ids_to_get}",
            headers=headers,
        )
        fields.append(json.loads(get_value_fields.content.decode("utf-8")))
    fields = flatten(fields)
    field_id_to_name_map = {str(value["id"]): value["name"] for value in fields}
    print("field id to name map", field_id_to_name_map)
    # get table id's from fields and repeat the process
    table_ids = set([item["scan_report_table"] for item in fields])
    paginated_table_ids = paginate(table_ids, 2000, "")
    tables = []
    for ids in paginated_table_ids:
        ids_to_get = ",".join(map(str, ids))

        get_field_tables = requests.get(
            url=f"{api_url}scanreporttablesfilter/?id__in={ids_to_get}",
            headers=headers,
        )
        tables.append(json.loads(get_field_tables.content.decode("utf-8")))
    tables = flatten(tables)

    # get all scan reports to be used to filter values by only values that come from active scan reports
    get_scan_reports = requests.get(
        url=f"{api_url}scanreports/",
        headers=headers,
    )
    # get active scanreports and map them to fields. Remove any fields in archived reports
    scanreports = json.loads(get_scan_reports.content.decode("utf-8"))
    active_reports = [
        str(item["id"])
        for item in scanreports
        if item["hidden"] == False and item["status"] == "COMPLET"
    ]
    # active reports is list of report ids that are not archived

    # map value id to active scan report
    table_id_to_active_scanreport_map = {
        str(element["id"]): str(element["scan_report"])
        for element in tables
        if str(element["scan_report"]) in active_reports
    }
    field_id_to_active_scanreport_map = {
        str(element["id"]): table_id_to_active_scanreport_map[
            str(element["scan_report_table"])
        ]
        for element in fields
        if str(element["scan_report_table"]) in table_id_to_active_scanreport_map
    }

    value_id_to_active_scanreport_map = {
        str(element["id"]): field_id_to_active_scanreport_map[
            str(element["scan_report_field"])
        ]
        for element in existing_scanreport_values
        if str(element["scan_report_field"]) in field_id_to_active_scanreport_map
    }
    existing_scanreport_values = [
        item
        for item in existing_scanreport_values
        if str(item["id"]) in value_id_to_active_scanreport_map
    ]
    print("VALUES TO SCAN REPORT", value_id_to_active_scanreport_map)
    print("FILTERED VALUES", existing_scanreport_values)

    #
    existing_mappings = [
        {
            "name": value["value"],
            "concept": value_id_to_concept_map[str(value["id"])],
            "id": value["id"],
            "description": value["value_description"],
            "field_name": field_id_to_name_map[str(value["scan_report_field"])],
        }
        for value in existing_scanreport_values
    ]

    value_name_to_id_map = {}
    for item in new_values_matching_list:
        name = item["name"]
        description = item["description"]
        field_name = item["field_name"]
        mappings_matching_value_name = [
            mapping
            for mapping in existing_mappings
            if mapping["name"] == name
            and mapping["description"] == description
            and mapping["field_name"] == field_name
        ]
        target_concept_ids = set(
            [mapping["concept"] for mapping in mappings_matching_value_name]
        )
        target_value_id = set(
            [mapping["id"] for mapping in mappings_matching_value_name]
        )
        if len(target_concept_ids) == 1:
            value_name_to_id_map[str(name)] = str(target_value_id.pop())

    concepts_to_post = []
    concept_response_content = []
    for name, id in new_value_ids.items():
        try:
            link_id = value_name_to_id_map[str(name)]
            print("VALUE link id", link_id)
            concept_id = value_id_to_concept_map[str(link_id)]
            print("VALUE concept id", value_name_to_id_map)
            # current_id=new_value_ids[str(name)]
            print(
                f"Found value with id: {link_id} with exsting concept mapping: {concept_id} which matches new value id: {id}"
            )
            # Create ScanReportConcept entry for copying over the concept
            concept_entry = {
                "nlp_entity": None,
                "nlp_entity_type": None,
                "nlp_confidence": None,
                "nlp_vocabulary": None,
                "nlp_processed_string": None,
                "concept": concept_id,
                "object_id": id,
                "content_type": content_type,
                "creation_type": "R",
            }
            concepts_to_post.append(concept_entry)

        except:
            continue
    if concepts_to_post:
        paginated_concepts_to_post = paginate(concepts_to_post)
        concept_response = []
        for concepts_to_post_item in paginated_concepts_to_post:
            get_concept_response = requests.post(
                url=api_url + "scanreportconcepts/",
                headers=headers,
                data=json.dumps(concepts_to_post_item),
            )
            print(
                "CONCEPTS SAVE STATUS >>>",
                get_concept_response.status_code,
                get_concept_response.reason,
                flush=True,
            )
            concept_response.append(
                json.loads(get_concept_response.content.decode("utf-8"))
            )
        concept_content = flatten(concept_response)

        concept_response_content += concept_content

        print("POST concepts all finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))


def remove_BOM(intermediate):
    return [
        {key.replace("\ufeff", ""): value for key, value in d.items()}
        for d in intermediate
    ]


def process_three_item_dict(three_item_data):
    """
    Converts a list of dictionaries (each with keys 'csv_file_name', 'field_name' and
    'code') to a nested dictionary with indices 'csv_file_name', 'field_name' and
    internal value 'code'.

    [{'csv_file_name': 'table1', 'field_name': 'field1', 'value': 'value1', 'code':
    'code1'},
    {'csv_file_name': 'table1', 'field_name': 'field2', 'value': 'value2'},
    {'csv_file_name': 'table2', 'field_name': 'field2', 'value': 'value2', 'code':
    'code2'},
    {'csv_file_name': 'table3', 'field_name': 'field3', 'value': 'value3', 'code':
    'code3'}]
    ->
    {'table1': {'field1': 'value1', 'field2': 'value2'},
     'table2': {'field2': 'value2'},
     'table3': {'field3': 'value3}
    }
    """
    csv_file_names = set(row["csv_file_name"] for row in three_item_data)

    # Initialise the dictionary with the keys, and each value set to a blank dict()
    new_vocab_dictionary = dict.fromkeys(csv_file_names, dict())

    # Fill each subdict with the data from the input list
    for row in three_item_data:
        new_vocab_dictionary[row["csv_file_name"]][row["field_name"]] = row["code"]

    return new_vocab_dictionary


def process_four_item_dict(four_item_data):
    """
    Converts a list of dictionaries (each with keys 'csv_file_name', 'field_name' and
    'code' and 'value') to a nested dictionary with indices 'csv_file_name',
    'field_name', 'code', and internal value 'value'.

    [{'csv_file_name': 'table1', 'field_name': 'field1', 'value': 'value1', 'code':
    'code1'},
    {'csv_file_name': 'table1', 'field_name': 'field2', 'value': 'value2', 'code':
    'code2'},
    {'csv_file_name': 'table2', 'field_name': 'field2', 'value': 'value2', 'code':
    'code2'},
    {'csv_file_name': 'table2', 'field_name': 'field2', 'value': 'value3', 'code':
    'code3'},
    {'csv_file_name': 'table3', 'field_name': 'field3', 'value': 'value3', 'code':
    'code3'}]
    ->
    {'table1': {'field1': {'value1': 'code1'}, 'field2': {'value2': 'code2'}},
     'table2': {'field2': {'value2': 'code2', 'value3': 'code3'}},
     'table3': {'field3': {'value3': 'code3'}}
    }
    """
    csv_file_names = set(row["csv_file_name"] for row in four_item_data)

    # Initialise the dictionary with the keys, and each value set to a blank dict()
    new_data_dictionary = dict.fromkeys(csv_file_names, dict())

    for row in four_item_data:
        if row["field_name"] not in new_data_dictionary[row["csv_file_name"]]:
            new_data_dictionary[row["csv_file_name"]][row["field_name"]] = dict()
        new_data_dictionary[row["csv_file_name"]][row["field_name"]][row["code"]] = row[
            "value"
        ]

    return new_data_dictionary


# @memory_profiler.profile(stream=profiler_logstream)
def parse_blobs(scan_report_blob, data_dictionary_blob):
    print("Get blobs", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
    # Set Storage Account connection string
    blob_service_client = BlobServiceClient.from_connection_string(
        os.environ.get("STORAGE_CONN_STRING")
    )

    # Grab scan report data from blob
    streamdownloader = (
        blob_service_client.get_container_client("scan-reports")
        .get_blob_client(scan_report_blob)
        .download_blob()
    )
    scanreport_bytes = BytesIO(streamdownloader.readall())
    wb = openpyxl.load_workbook(
        scanreport_bytes, data_only=True, keep_links=False, read_only=True
    )

    # If dictionary is present, also download dictionary
    if data_dictionary_blob != "None":
        # Access data as StorageStreamerDownloader class
        # Decode and split the stream using csv.reader()
        dict_client = blob_service_client.get_container_client("data-dictionaries")
        blob_dict_client = dict_client.get_blob_client(data_dictionary_blob)

        # Grab all rows with 4 elements for use as value descriptions
        data_dictionary_intermediate = list(
            row
            for row in csv.DictReader(
                blob_dict_client.download_blob().readall().decode("utf-8").splitlines()
            )
            if row["value"] != ""
        )
        # Remove BOM from start of file if it's supplied.
        dictionary_data = remove_BOM(data_dictionary_intermediate)

        # Convert to nested dictionaries, with structure
        # {tables: {fields: {values: value description}}}
        data_dictionary = process_four_item_dict(dictionary_data)

        # Grab all rows with 3 elements for use as possible vocabs
        vocab_dictionary_intermediate = list(
            row
            for row in csv.DictReader(
                blob_dict_client.download_blob().readall().decode("utf-8").splitlines()
            )
            if row["value"] == ""
        )
        vocab_data = remove_BOM(vocab_dictionary_intermediate)

        # Convert to nested dictionaries, with structure
        # {tables: {fields: vocab}}
        vocab_dictionary = process_three_item_dict(vocab_data)

    else:
        data_dictionary = None
        vocab_dictionary = None

    return wb, data_dictionary, vocab_dictionary


# @memory_profiler.profile(stream=profiler_logstream)
def post_tables(fo_ws, api_url, scan_report_id, headers):
    # Get all the table names in the order they appear in the Field Overview page
    table_names = []
    # Iterate over cells in the first column, but because we're in ReadOnly mode we
    # can't do that in the simplest manner.
    for row in fo_ws.iter_rows(min_row=2, max_row=fo_ws.max_row):
        cell_value = row[0].value
        # Check value is both non-empty and not seen before
        if cell_value and cell_value not in table_names:
            table_names.append(cell_value)

    """
    For each table create a scan_report_table entry,
    Append entry to table_entries_to_post[] list,
    Create JSON array with all the entries,
    Send POST request to API with JSON as input,
    Save the response data(table IDs)
    """
    table_entries_to_post = []
    # print("Working on Scan Report >>>", scan_report_id)
    print("RAM memory % used:", psutil.virtual_memory())
    print("TABLES NAMES >>> ", table_names)

    for table_name in table_names:
        # print("WORKING ON TABLE >>> ", table_name)

        # Truncate table names because sheet names are truncated to 31 characters in Excel
        short_table_name = table_name[:31]

        # Create ScanReportTable entry
        # Link to scan report using ID from the queue message
        table_entry = {
            "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
            "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
            "name": short_table_name,
            "scan_report": str(scan_report_id),
            "person_id": None,
            "birth_date": None,
            "measurement_date": None,
            "condition_date": None,
            "observation_date": None,
        }

        # print("SCAN REPORT TABLE ENTRY", table_entry)

        # Append to list
        table_entries_to_post.append(table_entry)

    print("POST tables", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
    # POST request to scanreporttables
    tables_response = requests.post(
        "{}scanreporttables/".format(api_url),
        data=json.dumps(table_entries_to_post),
        headers=headers,
    )

    print("POST tables finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))

    print("TABLE SAVE STATUS >>>", tables_response.status_code)
    # Error on failure
    if tables_response.status_code != 201:
        process_failure(api_url, scan_report_id, headers)
        raise HTTPError(
            " ".join(
                [
                    "Error in table save:",
                    str(tables_response.status_code),
                    str(json.dumps(table_entries_to_post)),
                ]
            )
        )
    print("RAM memory % used:", psutil.virtual_memory())

    # Load the result of the post request,
    tables_content = json.loads(tables_response.content.decode("utf-8"))

    # Save the table ids that were generated from the POST method
    table_ids = [element["id"] for element in tables_content]

    print("TABLE IDs", table_ids)
    table_name_to_id_map = dict(zip(table_names, table_ids))
    return table_name_to_id_map


# @memory_profiler.profile(stream=profiler_logstream)
async def process_values_from_sheet(
    sheet,
    data_dictionary,
    vocab_dictionary,
    current_table_name,
    names_to_ids_dict,
    api_url,
    scan_report_id,
    headers,
):
    # print("WORKING ON", sheet.title)
    # Reset list for values
    value_entries_to_post = []
    # Get (col_name, value, frequency) for each field in the table
    fieldname_value_freq_dict = process_scan_report_sheet_table(sheet)

    """
    For every result of process_scan_report_sheet_table,
    Save the current name,value,frequency
    Create ScanReportValue entry,
    Append to value_entries_to_post[] list,
    Create JSON array with all the value entries, 
    Send POST request to API with JSON as input
    """
    for name, value_freq_tuples in fieldname_value_freq_dict.items():
        for full_value, frequency in value_freq_tuples:
            value = full_value[0:127]

            if not frequency:
                frequency = 0

            if data_dictionary is not None:
                # Look up value description. We use .get() to guard against
                # nonexistence in the dictionary without having to manually check. It
                # returns None if the value is not present
                table = data_dictionary.get(
                    str(current_table_name)
                )  # dict of fields in table
                if table:
                    field = data_dictionary[str(current_table_name)].get(
                        str(name)
                    )  # dict of values in field in table
                    if field:
                        val_desc = data_dictionary[str(current_table_name)][
                            str(name)
                        ].get(str(value))
                    else:
                        val_desc = None
                else:
                    val_desc = None

                # Grab data from the 'code' column in the data dictionary
                # 'code' can contain an ordinary value (e.g. Yes, No, Nurse, Doctor)
                # or it could contain one of our pre-defined vocab names
                # e.g. SNOMED, RxNorm, ICD9 etc.
                # We use .get() to guard against nonexistence in the dictionary
                # without having to manually check. It returns None if the value is
                # not present
                table = vocab_dictionary.get(
                    str(current_table_name)
                )  # dict of fields in table
                if table:
                    code = vocab_dictionary[str(current_table_name)].get(
                        str(name)
                    )  # dict of values, will default to None if field not found in table
                else:
                    code = None

                # If 'code' is in our vocab list, try and convert the ScanReportValue
                # (concept code) to conceptID
                # If there's a faulty concept code for the vocab, fail gracefully and
                # set concept_id to default (-1)
                if code in vocabs:
                    try:
                        concept_id = omop_helpers.get_concept_from_concept_code(
                            concept_code=value,
                            vocabulary_id=code,
                            no_source_concept=True,
                        )
                        concept_id = concept_id["concept_id"]
                    except:
                        concept_id = -1
                else:
                    concept_id = -1

            else:
                val_desc = None
                concept_id = -1

            # Create ScanReportValue entry
            # We temporarily utilise the redundant 'conceptID' field in ScanReportValue
            # to save any looked up conceptIDs in the previous block of code.
            # The conceptID will be cleared later
            scan_report_value_entry = {
                "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "value": value,
                "frequency": int(frequency),
                "conceptID": concept_id,
                "value_description": val_desc,
                "scan_report_field": names_to_ids_dict[name],
            }

            # Append to list
            value_entries_to_post.append(scan_report_value_entry)

    print(
        "POST",
        len(value_entries_to_post),
        "values to table",
        current_table_name,
        datetime.utcnow().strftime("%H:%M:%S.%fZ"),
    )
    print("RAM memory % used:", psutil.virtual_memory())
    chunked_value_entries_to_post = perform_chunking(value_entries_to_post)
    values_response_content = []
    print("chunked values list len:", len(chunked_value_entries_to_post))
    timeout = httpx.Timeout(60.0, connect=30.0)

    page_count = 0
    for chunk in chunked_value_entries_to_post:
        async with httpx.AsyncClient(timeout=timeout) as client:
            tasks = []
            page_lengths = []
            for page in chunk:
                # POST value_entries_to_post to ScanReportValues model
                tasks.append(
                    asyncio.ensure_future(
                        client.post(
                            url="{}scanreportvalues/".format(api_url),
                            data=json.dumps(page),
                            headers=headers,
                        )
                    )
                )
                page_lengths.append(len(page))
                page_count += 1

            values_responses = await asyncio.gather(*tasks)

        for i, values_response in enumerate(values_responses):
            print(
                "VALUES SAVE STATUSES >>>",
                values_response.status_code,
                values_response.reason_phrase,
                page_lengths[i],
                flush=True,
            )
            if values_response.status_code != 201:
                process_failure(api_url, scan_report_id, headers)
                raise HTTPError(
                    " ".join(
                        [
                            "Error in values save:",
                            str(values_response.status_code),
                            str(json.dumps(page)),
                        ]
                    )
                )

            values_response_content += json.loads(
                values_response.content.decode("utf-8")
            )

    print("POST values all finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
    print("RAM memory % used:", psutil.virtual_memory())
    # Process conceptIDs in ScanReportValues
    # GET values where the conceptID != -1 (i.e. we've converted a concept code to conceptID in the previous code)
    print("GET posted values", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
    get_ids_of_posted_values = requests.get(
        url=api_url + "scanreportvaluepks/?scan_report=" + str(scan_report_id),
        headers=headers,
    )
    print("GET posted values finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))

    ids_of_posted_values = json.loads(get_ids_of_posted_values.content.decode("utf-8"))

    # Create a list for a bulk data upload to the ScanReportConcept model

    concept_id_data = [
        {
            "nlp_entity": None,
            "nlp_entity_type": None,
            "nlp_confidence": None,
            "nlp_vocabulary": None,
            "nlp_processed_string": None,
            "concept": concept["conceptID"],
            "object_id": concept["id"],
            # TODO: we should query this value from the API
            # - via ORM it would be ContentType.objects.get(model='scanreportvalue').id,
            # but that's not available from an Azure Function.
            "content_type": 17,
        }
        for concept in ids_of_posted_values
    ]

    print(
        "POST",
        len(concept_id_data),
        "concepts",
        datetime.utcnow().strftime("%H:%M:%S.%fZ"),
    )

    paginated_concept_id_data = paginate(concept_id_data)

    concepts_response_content = []

    for page in paginated_concept_id_data:

        # POST the ScanReportConcept data to the model
        concepts_response = requests.post(
            url=api_url + "scanreportconcepts/",
            headers=headers,
            data=json.dumps(page),
        )

        print(
            "CONCEPT SAVE STATUS >>> ",
            concepts_response.status_code,
            concepts_response.reason,
            flush=True,
        )
        if concepts_response.status_code != 201:
            process_failure(api_url, scan_report_id, headers)
            raise HTTPError(
                " ".join(
                    [
                        "Error in concept save:",
                        str(concepts_response.status_code),
                        str(json.dumps(page)),
                    ]
                )
            )

        concepts_content = json.loads(concepts_response.content.decode("utf-8"))
        concepts_response_content += concepts_content

    print("POST concepts all finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
    print("RAM memory % used:", psutil.virtual_memory())
    # Update ScanReportValue to remove any data added to the conceptID field
    # conceptID field only used temporarily to hold the converted concept code -> conceptID
    # Now the conceptID is saved to the correct model (ScanReportConcept) there's no
    # need for the concept ID to also be saved to ScanReportValue::conceptID

    # Reset conceptID to -1 (default). This doesn't need pagination because it's a
    # loop over all relevant fields anyway
    put_update_json = json.dumps({"conceptID": -1})

    print(
        "PATCH",
        len(ids_of_posted_values),
        "values",
        datetime.utcnow().strftime("%H:%M:%S.%fZ"),
    )
    for concept in ids_of_posted_values:
        print("PATCH value", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
        value_response = requests.patch(
            url=api_url + "scanreportvalues/" + str(concept["id"]) + "/",
            headers=headers,
            data=put_update_json,
        )
        # print("PATCH value finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
        if value_response.status_code != 200:
            process_failure(api_url, scan_report_id, headers)
            raise HTTPError(
                " ".join(
                    [
                        "Error in value save:",
                        str(value_response.status_code),
                        str(put_update_json),
                    ]
                )
            )

    print("PATCH values finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
    reuse_existing_field_concepts(names_to_ids_dict, 15, api_url, headers)
    reuse_existing_value_concepts(values_response_content, 17, api_url, headers)
    print("RAM memory % used:", psutil.virtual_memory())


def post_field_entries(field_entries_to_post, api_url, scan_report_id, headers):
    paginated_field_entries_to_post = paginate(field_entries_to_post)
    fields_response_content = []
    # POST Fields
    for page in paginated_field_entries_to_post:
        fields_response = requests.post(
            "{}scanreportfields/".format(api_url),
            data=json.dumps(page),
            headers=headers,
        )
        # print('dumped:', json.dumps(page))
        print(
            "FIELDS SAVE STATUS >>>",
            fields_response.status_code,
            fields_response.reason,
            len(page),
            flush=True,
        )

        if fields_response.status_code != 201:
            process_failure(api_url, scan_report_id, headers)
            raise HTTPError(
                " ".join(
                    [
                        "Error in fields save:",
                        str(fields_response.status_code),
                        str(json.dumps(page)),
                    ]
                )
            )

        fields_content = json.loads(fields_response.content.decode("utf-8"))
        # print('fc:',fields_content)
        fields_response_content += fields_content
        # print('frc:', fields_response_content)

    print(
        "POST fields all finished",
        datetime.utcnow().strftime("%H:%M:%S.%fZ"),
        flush=True,
    )
    return fields_response_content


def main(msg: func.QueueMessage):
    api_url, headers, scan_report_blob, data_dictionary_blob, scan_report_id = startup(
        msg
    )
    # Set the status to 'Upload in progress'
    status_in_progress_response = requests.patch(
        url=f"{api_url}scanreports/{scan_report_id}/",
        data=json.dumps({"status": "UPINPRO"}),
        headers=headers,
    )

    wb, data_dictionary, vocab_dictionary = parse_blobs(
        scan_report_blob, data_dictionary_blob
    )
    # Get the first sheet 'Field Overview',
    # to populate ScanReportTable & ScanReportField models
    fo_ws = wb.worksheets[0]

    table_name_to_id_map = post_tables(fo_ws, api_url, scan_report_id, headers)

    """
    POST fields per table:
    For each row in Field Overview create an entry for scan_report_field,
    Empty row signifies end of fields in a table
    Append field entry to field_entries_to_post[] list,
    Create JSON array with all the field entries, 
    Send POST request to API with JSON as input,
    Save the response data(field ids,field names) in a dictionary
    Set the current working sheet to be the same as the current table
    Post the values for that table
    """
    field_entries_to_post = []

    # Loop over all rows in Field Overview sheet.
    # For sheets past the first two in the Scan Report
    # i.e. all 'data' sheets that are not Field Overview and Table Overview
    print("Start fields loop", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
    previous_row_value = None
    for i, row in enumerate(fo_ws.iter_rows(min_row=2, max_row=fo_ws.max_row), start=2):
        # Guard against unnecessary rows beyond the last true row with contents
        if (previous_row_value is None or previous_row_value == "") and (
            row[0].value is None or row[0].value == ""
        ):
            break
        previous_row_value = row[0].value

        if row[0].value != "" and row[0].value is not None:
            current_table_name = row[0].value
            # Create ScanReportField entry
            field_entry = {
                "scan_report_table": table_name_to_id_map[current_table_name],
                "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "name": str(row[1].value),
                "description_column": str(row[2].value),
                "type_column": str(row[3].value),
                "max_length": row[4].value,
                "nrows": row[5].value,
                "nrows_checked": row[6].value,
                "fraction_empty": round(default_zero(row[7].value), 2),
                "nunique_values": row[8].value,
                "fraction_unique": round(default_zero(row[9].value), 2),
                "ignore_column": None,
                # "is_patient_id": False,
                # "is_ignore": False,
                # "pass_from_source": True,
                # "classification_system": str(row[11].value),
                # "concept_id": -1,
                # "field_description": None,
            }
            # Append each entry to a list
            field_entries_to_post.append(field_entry)

        else:
            # This is the scenario where the line is empty, so we're at the end of
            # the table. Don't add a field entry, but process all those so far.
            # print("scan_report_field_entries >>>", field_entries_to_post)

            # POST fields in this table
            print(
                "POST",
                len(field_entries_to_post),
                "fields to table",
                current_table_name,
                datetime.utcnow().strftime("%H:%M:%S.%fZ"),
            )
            print("RAM memory % used:", psutil.virtual_memory())

            fields_response_content = post_field_entries(
                field_entries_to_post, api_url, scan_report_id, headers
            )
            field_entries_to_post = []

            # Create a dictionary with field names and field ids from the response
            # as key value pairs
            # e.g ("Field Name": Field ID)
            names_to_ids_dict = {
                str(element.get("name", None)): str(element.get("id", None))
                for element in fields_response_content
            }

            # print("Dictionary id:name", names_to_ids_dict)

            if current_table_name not in wb.sheetnames:
                process_failure(api_url, scan_report_id, headers)
                raise ValueError(
                    f"Attempting to access sheet '{current_table_name}'"
                    f" in scan report, but no such sheet exists."
                )

            # Go to Table sheet to process all the values from the sheet
            sheet = wb[current_table_name]
            asyncio.run(
                process_values_from_sheet(
                    sheet,
                    data_dictionary,
                    vocab_dictionary,
                    current_table_name,
                    names_to_ids_dict,
                    api_url,
                    scan_report_id,
                    headers,
                )
            )

    # Set the status to 'Upload Complete'
    status_complete_response = requests.patch(
        url=f"{api_url}scanreports/{scan_report_id}/",
        data=json.dumps({"status": "UPCOMPL"}),
        headers=headers,
    )
    wb.close()
